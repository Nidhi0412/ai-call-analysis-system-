import os
import logging
import asyncio
from pathlib import Path
from typing import Dict, List, Optional, Tuple
import json
import time
from concurrent.futures import ThreadPoolExecutor, as_completed
import multiprocessing
import numpy as np
import librosa

# Local Whisper for transcription
import whisper

# Language detection and translation
from googletrans import Translator
from langdetect import detect, DetectorFactory

# Data handling
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

# Advanced speaker diarization
try:
    from Call_recordings_AI.unified_audio_processor import UnifiedAudioProcessor
except ImportError:
    from unified_audio_processor import UnifiedAudioProcessor

# Common infrastructure imports
import sys
sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

try:
    from config import CONFIG
except ImportError:
    # Fallback config if not available
    CONFIG = {
        'Development': {
            'APP_HOST': '0.0.0.0',
            'APP_PORT': 4000,
            'WORKERS': 1
        }
    }

try:
    from pylogger import pylogger
except ImportError:
    # Mock logger for testing
    class MockLogger:
        def __init__(self, *args, **kwargs):
            pass
        def log_it(self, data): 
            print(f"LOG: {data}")
    pylogger = MockLogger

# Configure logging using pylogger
app_env = os.getenv('NODE_ENV', 'Development')
env_config = CONFIG.get(app_env, {})
logger = pylogger("/home/saas/logs/innex", "call_recordings_ai")

# Set seed for consistent language detection
DetectorFactory.seed = 0

class TranscriptionWithSpeakersService:
    """
    Handles audio transcription with speaker identification using OpenAI Whisper
    """
    
    def __init__(self, api_key: str = None, model: str = "large-v2", use_advanced_diarization: bool = False, use_hinglish: bool = False, use_xglish: bool = False):
        """
        Initialize the transcription service with speaker identification
        
        Args:
            api_key: OpenAI API key (if None, will use environment variable)
            model: Whisper model to use (tiny, base, small, medium, large, large-v2, large-v3, large-v3-turbo)
            use_advanced_diarization: Whether to use advanced ML-based diarization
            use_hinglish: Whether to transcribe to Hinglish (Hindi in English script) instead of pure Hindi
            use_xglish: Whether to transcribe Indian languages to English script (Telugish, Tamglish, etc.)
        """
        # Initialize Local Whisper model
        self.model_name = model
        self.whisper_model = whisper.load_model(model)
        self.use_advanced_diarization = use_advanced_diarization
        self.use_hinglish = use_hinglish  # New parameter for Hinglish mode
        self.use_xglish = use_xglish  # New parameter for X-glish mode (Telugish, Tamglish, etc.)
        
        # Indian languages mapping for X-glish support
        self.indian_languages = {
            'hi': {'name': 'Hindi', 'english_mix': 'Hinglish', 'script': 'devanagari'},
            'te': {'name': 'Telugu', 'english_mix': 'Telugish', 'script': 'telugu'},
            'ta': {'name': 'Tamil', 'english_mix': 'Tamglish', 'script': 'tamil'},
            'kn': {'name': 'Kannada', 'english_mix': 'Kannadish', 'script': 'kannada'},
            'ml': {'name': 'Malayalam', 'english_mix': 'Malayglish', 'script': 'malayalam'},
            'bn': {'name': 'Bengali', 'english_mix': 'Benglish', 'script': 'bengali'},
            'mr': {'name': 'Marathi', 'english_mix': 'Maringlish', 'script': 'devanagari'},
            'gu': {'name': 'Gujarati', 'english_mix': 'Gujlish', 'script': 'gujarati'},
            'pa': {'name': 'Punjabi', 'english_mix': 'Punglish', 'script': 'gurmukhi'},
            'or': {'name': 'Odia', 'english_mix': 'Odiaglish', 'script': 'odia'},
            'as': {'name': 'Assamese', 'english_mix': 'Asglish', 'script': 'assamese'}
        }
        
        # Initialize translation service
        try:
            self.translator = Translator()
        except Exception as e:
            logger.log_it({
                "logType": "warning",
                "prefix": "transcription_with_speakers",
                "logData": {
                    "message": "Translation service initialization failed",
                    "error": str(e)
                }
            })
            self.translator = None
        
        # Initialize advanced diarization if requested
        self.advanced_diarization = None
        if use_advanced_diarization:
            try:
                # Initialize the advanced ML-based diarization system
                self.advanced_diarization = UnifiedAudioProcessor(n_speakers=2)
                logger.log_it({
                    "logType": "info",
                    "prefix": "transcription_with_speakers",
                    "logData": {
                        "message": "Advanced diarization system initialized",
                        "n_speakers": 2,
                        "features": "Voice fingerprinting, acoustic analysis, ML clustering"
                    }
                })
            except Exception as e:
                logger.log_it({
                    "logType": "warning",
                    "prefix": "transcription_with_speakers",
                    "logData": {
                        "message": "Advanced diarization initialization failed, using enhanced analysis",
                        "error": str(e)
                    }
                })
        
        # Token usage tracking
        self.token_usage = {
            "transcription_tokens": 0,
            "translation_tokens": 0,
            "total_tokens": 0
        }
        
        # Enhanced language mapping for better detection
        self.language_map = {
            'hi': 'Hindi',
            'hin': 'Hindi',
            'hindi': 'Hindi',
            'gu': 'Gujarati', 
            'en': 'English',
            'mr': 'Marathi',
            'bn': 'Bengali',
            'ta': 'Tamil',
            'te': 'Telugu',
            'kn': 'Kannada',
            'ml': 'Malayalam',
            'pa': 'Punjabi'
        }
        
        # Hindi-specific settings
        self.hindi_optimizations = {
            "force_hindi_detection": True,  # Try forcing Hindi for Hindi audio
            "enhanced_prompt": True,  # Use enhanced prompts for Hindi
            "segment_combining": True,  # Combine short segments for better context
            "min_segment_length": 2.0,  # Shorter segments for Hindi
            "max_segment_length": 20.0  # Longer segments for Hindi context
        }
        
        # Translation quality settings - More lenient for better coverage
        self.min_segment_length = 1.0  # Reduced from 3.0 to 1.0 for better coverage
        self.skip_short_segments = False  # Disabled to include all segments
        self.include_all_segments = True  # New flag to include all segments
        
        logger.log_it({
            "logType": "info",
            "prefix": "transcription_with_speakers",
            "logData": {
                "message": f"TranscriptionWithSpeakersService initialized",
                "model": model,
                "advanced_diarization": self.use_advanced_diarization,
                "use_hinglish_setting": self.use_hinglish,
                "note": "Automatic language detection will override use_hinglish setting"
            }
        })
    
    def transcribe_with_speakers(self, audio_path: str, language_override: str = None) -> Dict:
        """
        Transcribe audio file with speaker identification using OpenAI Whisper with Indian code-mixing support
        
        Args:
            audio_path: Path to the audio file
            
        Returns:
            dict: Transcription results including text, speakers, and timestamps
        """
        try:
            logger.log_it({
                "logType": "info",
                "prefix": "transcription_with_speakers",
                "logData": {
                    "message": f"Starting transcription with speakers for: {audio_path}",
                    "use_hinglish": self.use_hinglish
                }
            })
            
            # Check if file exists
            if not os.path.exists(audio_path):
                raise FileNotFoundError(f"Audio file not found: {audio_path}")
            
            # Enhanced prompt for Indian code-mixing with company-specific keywords
            indian_code_mixing_prompt = """
This is a customer service call in India with:
- Hindi as primary language with regional accent (Telugu/Bengali/Marathi/Punjabi/Gujarati)
- English words and phrases mixed in naturally
- Code-switching between Hindi, regional language, and English
- Customer service context with technical terms
- Regional pronunciation variations

IMPORTANT: If X-glish mode is enabled, transcribe Indian languages using English script:
- Telugu → Telugish: "namaskaram" instead of "నమస్కారం"
- Tamil → Tamglish: "vanakkam" instead of "வணக்கம்"  
- Kannada → Kannadish: "namaskara" instead of "ನಮಸ್ಕಾರ"
- Malayalam → Malayglish: "namaskaram" instead of "നമസ്കാരം"
- Bengali → Benglish: "namaskar" instead of "নমস্কার"
- Marathi → Maringlish: "namaskar" instead of "नमस्कार"
- Gujarati → Gujlish: "namaste" instead of "નમસ્તે"
- Punjabi → Punglish: "sat sri akal" instead of "ਸਤ ਸ੍ਰੀ ਅਕਾਲ"
- Odia → Odiaglish: "namaskar" instead of "ନମସ୍କାର"
- Assamese → Asglish: "namaskar" instead of "নমস্কাৰ"

IMPORTANT COMPANY-SPECIFIC KEYWORDS TO RECOGNIZE ACCURATELY:
- Company Names: Yanolja, Yanolja Cloud Solution, Ezee, Ultra Viewer, AnyDesk
- Technical Terms: Cloud Solution,Ultra Viewer, Remote Desktop, AnyDesk, TeamViewer
- Email/Contact: Gmail, Outlook, Yahoo, Hotmail, Mail ID, Email ID
- Common Phrases: "Yanolja Cloud Solution", "Ezee software", "Ultra Viewer", "AnyDesk connection"
- Business Terms: Booking, Reservation, Hotel, Restaurant, Room Charge, Billing
- Technical Support: Remote access, Screen sharing, Connection issues, Login problems
- Hotel Business Terms: Revenue, CityLedger, Front Desk, Check-in, Check-out, Room Rate, Occupancy, RevPAR, ADR, Hotel Management, PMS (Property Management System)
- Financial Terms: Invoice, Payment, Credit Card, Debit Card, Cash, Receipt, Tax, GST, Service Charge, Cancellation Fee, No-show Fee
- Hotel Operations: Housekeeping, Maintenance, Concierge, Bell Boy, Room Service, Housekeeping, Laundry, Mini Bar, Room Key, Key Card
- Revenue Management: Rate Management, Dynamic Pricing, Yield Management, Revenue Optimization, Market Segmentation, Demand Forecasting
- CityLedger Terms: City Ledger, Corporate Account, Direct Billing, Credit Limit, Payment Terms, Aging Report, Outstanding Balance, Credit Application
- Common Agent Requests: Hotel Code, Property Code, AnyDesk Number, Ultra Viewer Password, Mail ID Password, Login Credentials, Access Code, System Password
- Number Recognition Patterns: 
  * Direct numbers: "1, 2, 7, 8" or "one two seven eight"
  * Spelled out: "double nine seven four zero" = "99740"
  * Mixed format: "nine double seven four" = "9774"
  * Repeated digits: "double zero" = "00", "triple five" = "555"
  * Common patterns: "zero zero" = "00", "one one" = "11", "double zero" = "00"
- Authentication Terms: Username, Password, Login, Credentials, Access Code, PIN, OTP, Verification Code, Security Code

Please transcribe accurately maintaining:
- Mixed language structure in English script
- Regional accent patterns
- English technical terms
- Natural code-switching
- Customer service terminology
- Proper transliteration of Hindi/regional words
- Accurate recognition of company names and technical terms
- Proper spelling of email addresses and technical terms
"""
            
            # First, do a quick language detection to determine transcription approach
            logger.log_it({
                "logType": "info",
                "prefix": "transcription_with_speakers",
                "logData": "Performing initial language detection with Indian code-mixing support"
            })
            
            # Do initial transcription to detect language with timeout
            # Use neutral prompt for accurate language detection
            neutral_detection_prompt = """
This is an audio recording that may contain:
- English speech
- Hindi speech  
- Other Indian languages (Telugu, Tamil, Kannada, Bengali, Marathi, Gujarati, Punjabi, Odia, Assamese, Malayalam)
- Mixed language conversations
- Customer service calls
- Technical discussions

Please detect the primary language accurately without bias.
"""
            
            try:
                # Use local Whisper for transcription with detailed prompt
                # Use the EXACT same prompt as OpenAI API for fair comparison
                local_prompt = """This is a customer service call in India with:
- Hindi as primary language with regional accent (Telugu/Bengali/Marathi/Punjabi/Gujarati)
- English words and phrases mixed in naturally
- Code-switching between Hindi, regional language, and English
- Customer service context with technical terms
- Regional pronunciation variations

IMPORTANT COMPANY-SPECIFIC KEYWORDS TO RECOGNIZE ACCURATELY:
- Company Names: Yanolja, Yanolja Cloud Solution, Ezee, Ultra Viewer, AnyDesk
- Technical Terms: Cloud Solution,Ultra Viewer, Remote Desktop, AnyDesk, TeamViewer
- Email/Contact: Gmail, Outlook, Yahoo, Hotmail, Mail ID, Email ID
- Common Phrases: "Yanolja Cloud Solution", "Ezee software", "Ultra Viewer", "AnyDesk connection"
- Business Terms: Booking, Reservation, Hotel, Restaurant, Room Charge, Billing
- Technical Support: Remote access, Screen sharing, Connection issues, Login problems
- Hotel Business Terms: Revenue, CityLedger, Front Desk, Check-in, Check-out, Room Rate, Occupancy, RevPAR, ADR, Hotel Management, PMS (Property Management System)
- Financial Terms: Invoice, Payment, Credit Card, Debit Card, Cash, Receipt, Tax, GST, Service Charge, Cancellation Fee, No-show Fee
- Hotel Operations: Housekeeping, Maintenance, Concierge, Bell Boy, Room Service, Housekeeping, Laundry, Mini Bar, Room Key, Key Card
- Revenue Management: Rate Management, Dynamic Pricing, Yield Management, Revenue Optimization, Market Segmentation, Demand Forecasting
- CityLedger Terms: City Ledger, Corporate Account, Direct Billing, Credit Limit, Payment Terms, Aging Report, Outstanding Balance, Credit Application
- Common Agent Requests: Hotel Code, Property Code, AnyDesk Number, Ultra Viewer Password, Mail ID Password, Login Credentials, Access Code, System Password
- Number Recognition Patterns: 
  * Direct numbers: "1, 2, 7, 8" or "one two seven eight"
  * Spelled out: "double nine seven four zero" = "99740"
  * Mixed format: "nine double seven four" = "9774"
  * Repeated digits: "double zero" = "00", "triple five" = "555"
  * Common patterns: "zero zero" = "00", "one one" = "11", "double zero" = "00"
- Authentication Terms: Username, Password, Login, Credentials, Access Code, PIN, OTP, Verification Code, Security Code

Please transcribe accurately maintaining:
- Mixed language structure in English script
- Regional accent patterns
- English technical terms
- Natural code-switching
- Customer service terminology
- Proper transliteration of Hindi/regional words
- Accurate recognition of company names and technical terms
- Proper spelling of email addresses and technical terms"""
                
                result = self.whisper_model.transcribe(
                    audio_path,
                    language=None,  # Auto-detect language (matches OpenAI API)
                    word_timestamps=True,  # Equivalent to timestamp_granularities=["segment"]
                    verbose=False,
                    initial_prompt=local_prompt,  # Equivalent to OpenAI's prompt parameter
                    temperature=0.0,  # Matches OpenAI API exactly
                    beam_size=5,  # OpenAI's default beam search
                    best_of=5,  # OpenAI's default best_of
                    patience=1.0,  # OpenAI's default patience
                    length_penalty=1.0,  # OpenAI's default length penalty
                    suppress_tokens=[-1],  # Suppress special tokens like OpenAI
                    fp16=False  # Use FP32 for better accuracy (OpenAI uses FP32)
                )
                initial_transcript = {
                    "text": result["text"],
                    "language": result.get("language", "unknown"),
                    "segments": result.get("segments", [])
                }
            except Exception as e:
                logger.log_it({
                    "logType": "error",
                    "prefix": "transcription_with_speakers",
                    "logData": {
                        "message": "Initial transcription failed",
                        "error": str(e),
                        "file": audio_path
                    }
                })
                return {
                    "file_path": audio_path,
                    "transcribed_text": "",
                    "detected_language": "unknown",
                    "language_name": "Unknown",
                    "duration": 0,
                    "segments": [],
                    "speaker_stats": {},
                    "status": "error",
                    "error": f"Initial transcription failed: {str(e)}"
                }
            
            detected_language = initial_transcript["language"]
            
            # Extract transcript text for validation (used in both branches)
            transcript_text = initial_transcript["text"].lower() if initial_transcript["text"] else ""
            
            # Use language override if provided
            if language_override and language_override.strip():
                logger.log_it({
                    "logType": "info",
                    "prefix": "transcription_with_speakers",
                    "logData": {
                        "message": "Using language override",
                        "original_detection": detected_language,
                        "override": language_override
                    }
                })
                detected_language = language_override.strip()
            else:
                # Additional validation for language detection
                # Check if the detected language makes sense based on the transcript text
                
                # If detected as Indian language but text contains mostly English words, re-evaluate
                if detected_language in self.indian_languages and transcript_text:
                    english_word_count = sum(1 for word in transcript_text.split() if word.isalpha() and len(word) > 2)
                    total_word_count = len([word for word in transcript_text.split() if word.isalpha() and len(word) > 2])
                    
                    if total_word_count > 0:
                        english_ratio = english_word_count / total_word_count
                        
                        # If more than 70% of words are English, likely misdetected
                        if english_ratio > 0.7:
                            logger.log_it({
                                "logType": "info",
                                "prefix": "transcription_with_speakers",
                                "logData": {
                                    "message": "Language detection correction",
                                    "original_detection": detected_language,
                                    "english_ratio": english_ratio,
                                    "corrected_to": "en"
                                }
                            })
                            detected_language = "en"
            
            logger.log_it({
                "logType": "info",
                "prefix": "transcription_with_speakers",
                "logData": {
                    "message": "Language detected with validation",
                    "detected_language": detected_language,
                    "use_hinglish_setting": self.use_hinglish,
                    "use_xglish_setting": self.use_xglish,
                    "transcript_preview": transcript_text[:100] if transcript_text else "No text"
                }
            })
            
            # Determine transcription approach based on detected language
            logger.log_it({
                "logType": "info",
                "prefix": "transcription_with_speakers",
                "logData": {
                    "message": "Language detection debug",
                    "detected_language": detected_language,
                    "use_xglish": self.use_xglish,
                    "use_hinglish": self.use_hinglish,
                    "indian_languages_keys": list(self.indian_languages.keys())
                }
            })
            
            if detected_language in ["hi", "hin", "hindi"]:
                # Hindi detected - use Hinglish transcription with enhanced prompts
                transcription_language = "en"  # Use English for Hinglish
                transcription_method = "auto_hinglish_indian_code_mixing"
                auto_hinglish_enabled = True
                logger.log_it({
                    "logType": "info",
                    "prefix": "transcription_with_speakers",
                    "logData": "Hindi detected - automatically using Hinglish transcription with Indian code-mixing"
                })
            elif detected_language in self.indian_languages:
                # Other Indian language detected - use X-glish if enabled
                lang_info = self.indian_languages[detected_language]
                if self.use_xglish:
                    transcription_language = "en"  # Use English for X-glish
                    transcription_method = f"auto_{lang_info['english_mix'].lower()}_indian_code_mixing"
                    auto_hinglish_enabled = True
                    logger.log_it({
                        "logType": "info",
                        "prefix": "transcription_with_speakers",
                        "logData": f"{lang_info['name']} detected - using {lang_info['english_mix']} transcription with Indian code-mixing"
                    })
                else:
                    transcription_language = detected_language
                    transcription_method = f"standard_{detected_language}"
                    auto_hinglish_enabled = False
                    logger.log_it({
                        "logType": "info",
                        "prefix": "transcription_with_speakers",
                        "logData": f"{lang_info['name']} detected - using original script transcription"
                    })
            elif detected_language in ["en", "english"]:
                # English detected - use English transcription
                transcription_language = "en"
                transcription_method = "english"
                auto_hinglish_enabled = False
                logger.log_it({
                    "logType": "info",
                    "prefix": "transcription_with_speakers",
                    "logData": "English detected - using English transcription"
                })
            else:
                # Other language - use user's preference or default with Indian code-mixing support
                if self.use_hinglish or self.use_xglish:
                    transcription_language = "en"
                    transcription_method = "user_hinglish_indian_code_mixing"
                    auto_hinglish_enabled = True
                else:
                    transcription_language = None  # Auto-detect
                    transcription_method = "standard"
                    auto_hinglish_enabled = False
                logger.log_it({
                    "logType": "info",
                    "prefix": "transcription_with_speakers",
                    "logData": f"Other language detected ({detected_language}) - using {transcription_method}"
                })
            
            # Create dynamic prompt based on language and X-glish setting
            if auto_hinglish_enabled and detected_language in self.indian_languages:
                lang_info = self.indian_languages[detected_language]
                logger.log_it({
                    "logType": "info",
                    "prefix": "transcription_with_speakers",
                    "logData": {
                        "message": f"Creating X-glish prompt for {lang_info['name']}",
                        "language": lang_info['name'],
                        "xglish_variant": lang_info['english_mix'],
                        "transcription_language": transcription_language
                    }
                })
                dynamic_prompt = f"""
This is a customer service call in India with:
- {lang_info['name']} as primary language with regional accent
- English words and phrases mixed in naturally
- Code-switching between {lang_info['name']}, regional language, and English
- Customer service context with technical terms
- Regional pronunciation variations

IMPORTANT: Transcribe {lang_info['name']} using English script ({lang_info['english_mix']}):
- Use English alphabet instead of {lang_info['script']} script
- Example: "{lang_info['name']} word" instead of "{lang_info['script']} script"
- Maintain pronunciation accuracy in English script
- Keep English words as-is

IMPORTANT COMPANY-SPECIFIC KEYWORDS TO RECOGNIZE ACCURATELY:
- Company Names: Yanolja, Yanolja Cloud Solution, Ezee, Ultra Viewer, AnyDesk
- Technical Terms: Cloud Solution, Ultra Viewer, Remote Desktop, AnyDesk, TeamViewer
- Email/Contact: Gmail, Outlook, Yahoo, Hotmail, Mail ID, Email ID
- Common Phrases: "Yanolja Cloud Solution", "Ezee software", "Ultra Viewer", "AnyDesk connection"
- Business Terms: Booking, Reservation, Hotel, Restaurant, Room Charge, Billing
- Technical Support: Remote access, Screen sharing, Connection issues, Login problems
- Hotel Business Terms: Revenue, CityLedger, Front Desk, Check-in, Check-out, Room Rate, Occupancy, RevPAR, ADR, Hotel Management, PMS (Property Management System)
- Financial Terms: Invoice, Payment, Credit Card, Debit Card, Cash, Receipt, Tax, GST, Service Charge, Cancellation Fee, No-show Fee
- Hotel Operations: Housekeeping, Maintenance, Concierge, Bell Boy, Room Service, Housekeeping, Laundry, Mini Bar, Room Key, Key Card
- Revenue Management: Rate Management, Dynamic Pricing, Yield Management, Revenue Optimization, Market Segmentation, Demand Forecasting
- CityLedger Terms: City Ledger, Corporate Account, Direct Billing, Credit Limit, Payment Terms, Aging Report, Outstanding Balance, Credit Application
- Common Agent Requests: Hotel Code, Property Code, AnyDesk Number, Ultra Viewer Password, Mail ID Password, Login Credentials, Access Code, System Password
- Number Recognition Patterns: 
  * Direct numbers: "1, 2, 7, 8" or "one two seven eight"
  * Spelled out: "double nine seven four zero" = "99740"
  * Mixed format: "nine double seven four" = "9774"
  * Indian number patterns: "double zero" = "00", "triple five" = "555"
- Regional Variations: Account for Telugu/Bengali/Marathi/Punjabi/Gujarati pronunciation of English words
- Code-mixing: Natural switching between {lang_info['name']} and English within sentences
- Context Awareness: Customer service, technical support, hotel operations, billing inquiries
- Pronunciation: Regional Indian accent variations in English words
- Technical Terms: Software names, system commands, error messages in English
- Business Context: Hotel management, reservation systems, customer support workflows
"""
            else:
                dynamic_prompt = indian_code_mixing_prompt
            
            # Now do the actual transcription with determined approach
            try:
                # Use local Whisper for transcription with detailed prompt
                # Use the EXACT same prompt as OpenAI API for fair comparison
                local_prompt = """This is a customer service call in India with:
- Hindi as primary language with regional accent (Telugu/Bengali/Marathi/Punjabi/Gujarati)
- English words and phrases mixed in naturally
- Code-switching between Hindi, regional language, and English
- Customer service context with technical terms
- Regional pronunciation variations

IMPORTANT COMPANY-SPECIFIC KEYWORDS TO RECOGNIZE ACCURATELY:
- Company Names: Yanolja, Yanolja Cloud Solution, Ezee, Ultra Viewer, AnyDesk
- Technical Terms: Cloud Solution,Ultra Viewer, Remote Desktop, AnyDesk, TeamViewer
- Email/Contact: Gmail, Outlook, Yahoo, Hotmail, Mail ID, Email ID
- Common Phrases: "Yanolja Cloud Solution", "Ezee software", "Ultra Viewer", "AnyDesk connection"
- Business Terms: Booking, Reservation, Hotel, Restaurant, Room Charge, Billing
- Technical Support: Remote access, Screen sharing, Connection issues, Login problems
- Hotel Business Terms: Revenue, CityLedger, Front Desk, Check-in, Check-out, Room Rate, Occupancy, RevPAR, ADR, Hotel Management, PMS (Property Management System)
- Financial Terms: Invoice, Payment, Credit Card, Debit Card, Cash, Receipt, Tax, GST, Service Charge, Cancellation Fee, No-show Fee
- Hotel Operations: Housekeeping, Maintenance, Concierge, Bell Boy, Room Service, Housekeeping, Laundry, Mini Bar, Room Key, Key Card
- Revenue Management: Rate Management, Dynamic Pricing, Yield Management, Revenue Optimization, Market Segmentation, Demand Forecasting
- CityLedger Terms: City Ledger, Corporate Account, Direct Billing, Credit Limit, Payment Terms, Aging Report, Outstanding Balance, Credit Application
- Common Agent Requests: Hotel Code, Property Code, AnyDesk Number, Ultra Viewer Password, Mail ID Password, Login Credentials, Access Code, System Password
- Number Recognition Patterns: 
  * Direct numbers: "1, 2, 7, 8" or "one two seven eight"
  * Spelled out: "double nine seven four zero" = "99740"
  * Mixed format: "nine double seven four" = "9774"
  * Repeated digits: "double zero" = "00", "triple five" = "555"
  * Common patterns: "zero zero" = "00", "one one" = "11", "double zero" = "00"
- Authentication Terms: Username, Password, Login, Credentials, Access Code, PIN, OTP, Verification Code, Security Code

Please transcribe accurately maintaining:
- Mixed language structure in English script
- Regional accent patterns
- English technical terms
- Natural code-switching
- Customer service terminology
- Proper transliteration of Hindi/regional words
- Accurate recognition of company names and technical terms
- Proper spelling of email addresses and technical terms"""
                
                result = self.whisper_model.transcribe(
                    audio_path,
                    language=transcription_language if transcription_language != "auto" else None,
                    word_timestamps=True,  # Equivalent to timestamp_granularities=["segment"]
                    verbose=False,
                    initial_prompt=local_prompt,  # Equivalent to OpenAI's prompt parameter
                    temperature=0.0,  # Matches OpenAI API exactly
                    beam_size=5,  # OpenAI's default beam search
                    best_of=5,  # OpenAI's default best_of
                    patience=1.0,  # OpenAI's default patience
                    length_penalty=1.0,  # OpenAI's default length penalty
                    suppress_tokens=[-1],  # Suppress special tokens like OpenAI
                    fp16=False  # Use FP32 for better accuracy (OpenAI uses FP32)
                )
                transcript = {
                    "text": result["text"],
                    "language": result.get("language", transcription_language),
                    "segments": result.get("segments", [])
                }
            except Exception as e:
                logger.log_it({
                    "logType": "error",
                    "prefix": "transcription_with_speakers",
                    "logData": {
                        "message": "Main transcription failed",
                        "error": str(e),
                        "file": audio_path
                    }
                })
                return {
                    "file_path": audio_path,
                    "transcribed_text": "",
                    "detected_language": detected_language,
                    "language_name": self.language_map.get(detected_language, detected_language.title() if detected_language else "Unknown"),
                    "duration": 0,
                    "segments": [],
                    "speaker_stats": {},
                    "status": "error",
                    "error": f"Main transcription failed: {str(e)}"
                }
            
            # Track transcription usage (Whisper doesn't return token usage, so we estimate)
            # Whisper models typically use ~1 token per 4 characters of audio
            estimated_tokens = len(transcript["text"]) // 4 if transcript["text"] else 0
            self.token_usage["transcription_tokens"] += estimated_tokens
            self.token_usage["total_tokens"] += estimated_tokens
            
            logger.log_it({
                "logType": "info",
                "prefix": "transcription_with_speakers",
                "logData": {
                    "message": "Transcription token usage tracked",
                    "estimated_tokens": estimated_tokens,
                    "text_length": len(transcript["text"]) if transcript["text"] else 0,
                    "transcription_method": transcription_method,
                    "auto_hinglish_enabled": auto_hinglish_enabled
                }
            })
            
            # Process segments to identify speakers
            processed_segments = self._process_speaker_segments(transcript["segments"], audio_path)
            
            # Extract results
            result = {
                "file_path": audio_path,
                "transcribed_text": transcript["text"],
                "detected_language": detected_language,
                "language_name": self.language_map.get(detected_language, detected_language.title() if detected_language else "Unknown"),
                "duration": transcript.get("duration", 0),
                "segments": processed_segments,
                "speaker_stats": self._calculate_speaker_stats(processed_segments),
                "status": "success",
                "transcription_method": transcription_method,
                "auto_hinglish_enabled": auto_hinglish_enabled
            }
            
            logger.log_it({
                "logType": "info",
                "prefix": "transcription_with_speakers",
                "logData": {
                    "message": "Transcription with speakers successful",
                    "language": detected_language,
                    "text_length": len(transcript["text"]) if transcript["text"] else 0,
                    "duration": transcript.get("duration", 0),
                    "speakers_detected": len(set(seg.get("speaker", "") for seg in processed_segments)),
                    "transcription_method": transcription_method,
                    "auto_hinglish_enabled": auto_hinglish_enabled
                }
            })
            
            return result
            
        except Exception as e:
            logger.log_it({
                "logType": "error",
                "prefix": "transcription_with_speakers",
                "logData": {
                    "message": "Transcription with speakers failed",
                    "file": audio_path,
                    "error": str(e)
                }
            })
            return {
                "file_path": audio_path,
                "transcribed_text": "",
                "detected_language": "unknown",
                "language_name": "Unknown",
                "duration": 0,
                "segments": [],
                "speaker_stats": {},
                "status": "error",
                "error": str(e)
            }
    
    def _process_speaker_segments(self, segments: List, audio_path: str = None) -> List[Dict]:
        """
        Process segments to identify speakers based on voice characteristics
        
        Args:
            segments: Raw segments from Whisper (TranscriptionSegment objects)
            audio_path: Path to audio file for advanced diarization
            
        Returns:
            List of segments with speaker identification
        """
        processed_segments = []
        
        # Use advanced diarization if available and audio path provided
        if self.use_advanced_diarization and self.advanced_diarization and audio_path:
            return self._process_speakers_with_advanced_diarization(segments, audio_path)
        
        # Enhanced speaker identification using voice characteristics
        speaker_characteristics = self._analyze_speaker_characteristics(segments)
        
        for i, segment in enumerate(segments):
            # Handle both TranscriptionSegment objects and dictionaries
            if hasattr(segment, 'start'):
                # TranscriptionSegment object
                start_time = segment.start
                end_time = segment.end
                text = segment.text
                confidence = getattr(segment, 'confidence', 1.0)
            else:
                # Dictionary (fallback)
                start_time = segment.get('start', 0)
                end_time = segment.get('end', 0)
                text = segment.get('text', '')
                confidence = segment.get('confidence', 1.0)
            
            # Determine speaker using enhanced analysis
            speaker = self._determine_speaker_enhanced(i, text, start_time, end_time, speaker_characteristics)
            
            processed_segment = {
                "start": start_time,
                "end": end_time,
                "text": text,
                "confidence": confidence,
                "speaker": speaker,
                "speaker_type": "Agent" if speaker == "Speaker 1" else "Caller",
                "segment_id": i + 1,
                "diarization_method": "enhanced_analysis"
            }
            
            processed_segments.append(processed_segment)
        
        return processed_segments
    
    def _process_speakers_with_advanced_diarization(self, segments: List, audio_path: str) -> List[Dict]:
        """
        Process speakers using advanced ML-based diarization with actual audio data
        
        Args:
            segments: Raw segments from Whisper
            audio_path: Path to audio file for feature extraction
            
        Returns:
            List of segments with advanced speaker identification
        """
        try:
            logger.log_it({
                "logType": "info",
                "prefix": "transcription_with_speakers",
                "logData": "Using advanced ML-based speaker diarization with audio data"
            })
            
            # Load audio for advanced diarization
            audio, sr = librosa.load(audio_path, sr=16000)
            
            # Extract segment info and create audio segments
            segment_info = []
            audio_segments = []
            
            for i, segment in enumerate(segments):
                if hasattr(segment, 'start'):
                    start_time = segment.start
                    end_time = segment.end
                    text = segment.text
                    confidence = getattr(segment, 'confidence', 1.0)
                else:
                    start_time = segment.get('start', 0)
                    end_time = segment.get('end', 0)
                    text = segment.get('text', '')
                    confidence = segment.get('confidence', 1.0)
                
                # Extract audio segment for this time window
                start_sample = int(start_time * sr)
                end_sample = int(end_time * sr)
                
                if start_sample < len(audio) and end_sample <= len(audio):
                    audio_segment = audio[start_sample:end_sample]
                    audio_segments.append(audio_segment)
                    
                    segment_info.append({
                        'index': i,
                        'start': start_time,
                        'end': end_time,
                        'text': text,
                        'confidence': confidence,
                        'audio_segment': audio_segment
                    })
                else:
                    # Fallback for invalid time ranges
                    segment_info.append({
                        'index': i,
                        'start': start_time,
                        'end': end_time,
                        'text': text,
                        'confidence': confidence,
                        'audio_segment': None
                    })
            
            # Extract features for all segments using advanced diarization
            features = []
            valid_segments = []
            
            for info in segment_info:
                if info['audio_segment'] is not None and len(info['audio_segment']) > 0:
                    try:
                        feature_vector = self.advanced_diarization.extract_speaker_features(info['audio_segment'])
                        features.append(feature_vector)
                        valid_segments.append(info)
                    except Exception as e:
                        logger.log_it({
                            "logType": "warning",
                            "prefix": "transcription_with_speakers",
                            "logData": {
                                "message": "Feature extraction failed for segment",
                                "segment_index": info['index'],
                                "error": str(e)
                            }
                        })
                        # Add fallback feature vector
                        features.append(np.zeros(100))  # Default feature vector
                        valid_segments.append(info)
                else:
                    # Add fallback for segments without audio
                    features.append(np.zeros(100))
                    valid_segments.append(info)
            
            # Cluster speakers using advanced diarization
            if len(features) > 0:
                speaker_labels = self.advanced_diarization._cluster_speakers(features)
                
                logger.log_it({
                    "logType": "info",
                    "prefix": "transcription_with_speakers",
                    "logData": {
                        "message": "Advanced diarization clustering completed",
                        "segments_processed": len(valid_segments),
                        "features_extracted": len(features),
                        "speaker_distribution": dict(zip(*np.unique(speaker_labels, return_counts=True)))
                    }
                })
            else:
                # Fallback to enhanced analysis if no features extracted
                logger.log_it({
                    "logType": "warning",
                    "prefix": "transcription_with_speakers",
                    "logData": "No features extracted, using enhanced analysis fallback"
                })
                speaker_characteristics = self._analyze_speaker_characteristics(segments)
                speaker_labels = [self._determine_speaker_enhanced(i, info['text'], info['start'], info['end'], speaker_characteristics) 
                                for i, info in enumerate(valid_segments)]
            
            # Create processed segments with advanced speaker labels
            processed_segments = []
            for i, info in enumerate(valid_segments):
                if i < len(speaker_labels):
                    speaker = speaker_labels[i]
                else:
                    speaker = f"Speaker {(i % 2) + 1}"  # Fallback
                
                processed_segment = {
                    "start": info['start'],
                    "end": info['end'],
                    "text": info['text'],
                    "confidence": info['confidence'],
                    "speaker": speaker,
                    "speaker_type": "Agent" if speaker == "Speaker 1" else "Caller",
                    "segment_id": info['index'] + 1,
                    "diarization_method": "advanced_ml"
                }
                
                processed_segments.append(processed_segment)
            
            logger.log_it({
                "logType": "info",
                "prefix": "transcription_with_speakers",
                "logData": {
                    "message": "Advanced diarization processing completed successfully",
                    "segments_processed": len(processed_segments),
                    "diarization_method": "advanced_ml"
                }
            })
            
            return processed_segments
            
        except Exception as e:
            logger.log_it({
                "logType": "error",
                "prefix": "transcription_with_speakers",
                "logData": {
                    "message": "Advanced diarization failed, using enhanced analysis",
                    "error": str(e)
                }
            })
            
            # Fallback to enhanced analysis
            speaker_characteristics = self._analyze_speaker_characteristics(segments)
            
            processed_segments = []
            for i, segment in enumerate(segments):
                if hasattr(segment, 'start'):
                    start_time = segment.start
                    end_time = segment.end
                    text = segment.text
                    confidence = getattr(segment, 'confidence', 1.0)
                else:
                    start_time = segment.get('start', 0)
                    end_time = segment.get('end', 0)
                    text = segment.get('text', '')
                    confidence = segment.get('confidence', 1.0)
                
                speaker = self._determine_speaker_enhanced(i, text, start_time, end_time, speaker_characteristics)
                
                processed_segment = {
                    "start": start_time,
                    "end": end_time,
                    "text": text,
                    "confidence": confidence,
                    "speaker": speaker,
                    "speaker_type": "Agent" if speaker == "Speaker 1" else "Caller",
                    "segment_id": i + 1,
                    "diarization_method": "enhanced_fallback_error"
                }
                
                processed_segments.append(processed_segment)
            
            return processed_segments
    
    def _analyze_speaker_characteristics(self, segments: List) -> Dict:
        """
        Analyze voice characteristics to identify different speakers
        
        Args:
            segments: Raw segments from Whisper
            
        Returns:
            dict: Speaker characteristics analysis
        """
        speaker_analysis = {
            "speaker_1_patterns": [],
            "speaker_2_patterns": [],
            "segment_lengths": [],
            "text_patterns": []
        }
        
        # Analyze segment patterns
        for segment in segments:
            if hasattr(segment, 'text'):
                text = segment.text
                duration = segment.end - segment.start
            else:
                text = segment.get('text', '')
                duration = segment.get('end', 0) - segment.get('start', 0)
            
            speaker_analysis["segment_lengths"].append(duration)
            speaker_analysis["text_patterns"].append({
                "text": text,
                "length": len(text),
                "word_count": len(text.split()),
                "duration": duration
            })
        
        # Identify patterns for speaker differentiation
        if len(speaker_analysis["segment_lengths"]) > 1:
            avg_duration = sum(speaker_analysis["segment_lengths"]) / len(speaker_analysis["segment_lengths"])
            speaker_analysis["avg_segment_duration"] = avg_duration
            
            # Analyze text patterns
            text_lengths = [p["length"] for p in speaker_analysis["text_patterns"]]
            word_counts = [p["word_count"] for p in speaker_analysis["text_patterns"]]
            
            speaker_analysis["avg_text_length"] = sum(text_lengths) / len(text_lengths) if text_lengths else 0
            speaker_analysis["avg_word_count"] = sum(word_counts) / len(word_counts) if word_counts else 0
        
        return speaker_analysis
    
    def _determine_speaker_enhanced(self, segment_index: int, text: str, start_time: float, 
                                  end_time: float, speaker_characteristics: Dict) -> str:
        """
        Enhanced speaker determination using multiple factors
        
        Args:
            segment_index: Index of the segment
            text: Text content
            start_time: Start time of segment
            end_time: End time of segment
            speaker_characteristics: Analysis of speaker patterns
            
        Returns:
            str: Speaker identifier
        """
        # Factor 1: Position in conversation
        if segment_index == 0:
            return "Speaker 1"  # First speaker is usually the agent
        
        # Factor 2: Segment duration analysis
        duration = end_time - start_time
        avg_duration = speaker_characteristics.get("avg_segment_duration", 0)
        
        # Factor 3: Text length analysis
        text_length = len(text)
        avg_text_length = speaker_characteristics.get("avg_text_length", 0)
        
        # Factor 4: Word count analysis
        word_count = len(text.split())
        avg_word_count = speaker_characteristics.get("avg_word_count", 0)
        
        # Factor 5: Timing patterns
        time_since_start = start_time
        
        # Enhanced decision logic
        speaker_score_1 = 0
        speaker_score_2 = 0
        
        # Agent typically has shorter, more professional responses
        if avg_duration > 0:
            if duration < avg_duration * 0.8:
                speaker_score_1 += 1  # Agent
            elif duration > avg_duration * 1.2:
                speaker_score_2 += 1  # Caller
        
        # Agent typically has more concise text
        if avg_text_length > 0:
            if text_length < avg_text_length * 0.8:
                speaker_score_1 += 1
            elif text_length > avg_text_length * 1.2:
                speaker_score_2 += 1
        
        # Agent typically uses fewer words
        if avg_word_count > 0:
            if word_count < avg_word_count * 0.8:
                speaker_score_1 += 1
            elif word_count > avg_word_count * 1.2:
                speaker_score_2 += 1
        
        # Early in conversation, more likely to be agent
        if time_since_start < 30:  # First 30 seconds
            speaker_score_1 += 1
        
        # Professional language indicators (agent)
        professional_words = ['thank', 'you', 'call', 'service', 'help', 'assist', 'please', 'welcome', 'good', 'morning', 'afternoon', 'evening', 'how', 'may', 'i', 'help']
        text_lower = text.lower()
        professional_count = sum(1 for word in professional_words if word in text_lower)
        if professional_count >= 2:
            speaker_score_1 += 1
        
        # Question indicators (caller)
        question_words = ['what', 'how', 'why', 'when', 'where', 'can', 'could', 'would', 'will', 'should', 'is', 'are', 'do', 'does']
        question_count = sum(1 for word in question_words if word in text_lower)
        if question_count >= 1:
            speaker_score_2 += 1
        
        # Greeting patterns (agent)
        greeting_words = ['hello', 'hi', 'good', 'morning', 'afternoon', 'evening', 'welcome', 'thank', 'goodbye', 'bye', 'have', 'nice', 'day','welcome to Ezee','how may i assist you']
        if any(word in text_lower for word in greeting_words) and time_since_start < 10:
            speaker_score_1 += 2  # Strong indicator for agent
        
        # Farewell patterns (agent)
        farewell_words = ['thank', 'goodbye', 'bye', 'have', 'nice', 'day']
        if any(word in text_lower for word in farewell_words) and time_since_start > 60:
            speaker_score_1 += 1
        
        # Determine speaker based on scores
        if speaker_score_1 > speaker_score_2:
            return "Speaker 1"  # Agent
        elif speaker_score_2 > speaker_score_1:
            return "Speaker 2"  # Caller
        else:
            # Fallback to alternating pattern with bias towards agent
            if segment_index % 3 == 0:  # Every 3rd segment
                return "Speaker 1"  # Agent
            else:
                return "Speaker 1" if segment_index % 2 == 0 else "Speaker 2"
    
    def _calculate_speaker_stats(self, segments: List[Dict]) -> Dict:
        """
        Calculate statistics for each speaker
        
        Args:
            segments: Processed segments with speaker information
            
        Returns:
            Dictionary with speaker statistics
        """
        speaker_stats = {
            "Speaker 1": {"count": 0, "total_duration": 0, "total_words": 0},
            "Speaker 2": {"count": 0, "total_duration": 0, "total_words": 0}
        }
        
        for segment in segments:
            # Handle both dictionary and object access
            if isinstance(segment, dict):
                speaker = segment.get('speaker', 'Speaker 1')
                start_time = segment.get('start', 0)
                end_time = segment.get('end', 0)
                text = segment.get('text', '')
            else:
                # Handle object attributes
                speaker = getattr(segment, 'speaker', 'Speaker 1')
                start_time = getattr(segment, 'start', 0)
                end_time = getattr(segment, 'end', 0)
                text = getattr(segment, 'text', '')
            
            duration = end_time - start_time
            words = len(text.split()) if text else 0
            
            if speaker in speaker_stats:
                speaker_stats[speaker]["count"] += 1
                speaker_stats[speaker]["total_duration"] += duration
                speaker_stats[speaker]["total_words"] += words
        
        return speaker_stats
    
    async def translate_to_english(self, text: str, source_lang: str = None) -> Dict:
        """
        Translate text to English (async version) with improved handling
        
        Args:
            text: Text to translate
            source_lang: Source language code (if None, will auto-detect)
            
        Returns:
            dict: Translation results
        """
        try:
            if not text.strip():
                return {
                    "original_text": text,
                    "translated_text": "",
                    "source_language": source_lang or "unknown",
                    "status": "empty_text"
                }
            
            # Skip translation for very short text (likely noise)
            if len(text.strip()) < 3:
                return {
                    "original_text": text,
                    "translated_text": text,
                    "source_language": source_lang or "unknown",
                    "status": "too_short"
                }
            
            # Auto-detect language if not provided
            if not source_lang:
                try:
                    source_lang = detect(text)
                except:
                    source_lang = "unknown"
            
            # Skip translation if already English
            if source_lang == "en":
                return {
                    "original_text": text,
                    "translated_text": text,
                    "source_language": "en",
                    "status": "already_english"
                }
            
            # Try translation with timeout
            try:
                translation = await asyncio.wait_for(
                    self.translator.translate(text, dest='en', src=source_lang),
                    timeout=10.0
                )
                
                # Handle translation result
                translated_text = text  # Default to original text
                
                if hasattr(translation, 'text'):
                    translated_text = translation.text
                else:
                    translated_text = str(translation)
                
                # Validate translation quality
                if len(translated_text.strip()) < 2 or translated_text == text:
                    # Translation failed or didn't change anything
                    return {
                        "original_text": text,
                        "translated_text": text,
                        "source_language": source_lang,
                        "status": "translation_failed"
                    }
                
                return {
                    "original_text": text,
                    "translated_text": translated_text,
                    "source_language": source_lang,
                    "status": "success"
                }
                
            except asyncio.TimeoutError:
                logger.log_it({
                    "logType": "warning",
                    "prefix": "transcription_with_speakers",
                    "logData": {
                        "message": "Translation timeout, using original text",
                        "text": text[:50] + "..." if len(text) > 50 else text
                    }
                })
                return {
                    "original_text": text,
                    "translated_text": text,
                    "source_language": source_lang,
                    "status": "timeout"
                }
            
        except Exception as e:
            logger.log_it({
                "logType": "error",
                "prefix": "transcription_with_speakers",
                "logData": {
                    "message": "Translation failed",
                    "error": str(e),
                    "text_length": len(text) if text else 0
                }
            })
            return {
                "original_text": text,
                "translated_text": text,  # Return original if translation fails
                "source_language": source_lang or "unknown",
                "status": "error",
                "error": str(e)
            }
    
    async def translate_with_openai(self, text: str, source_lang: str = None) -> Dict:
        """
        Translate text using OpenAI API (better quality than googletrans)
        
        Args:
            text: Text to translate
            source_lang: Source language code
            
        Returns:
            dict: Translation results
        """
        try:
            if not text.strip() or len(text.strip()) < 3:
                return {
                    "original_text": text,
                    "translated_text": text,
                    "source_language": source_lang or "unknown",
                    "status": "too_short"
                }
            
            # Create translation prompt
            language_name = self.language_map.get(source_lang, source_lang) if source_lang else "the detected language"
            prompt = f"Translate the following {language_name} text to English. Provide only the English translation:\n\n{text}"
            
            # Use OpenAI for translation with timeout
            try:
                response = await asyncio.wait_for(
                    asyncio.get_event_loop().run_in_executor(
                        None,
                        lambda: self.client.chat.completions.create(
                            model="gpt-3.5-turbo",
                            messages=[
                                {"role": "system", "content": "You are a professional translator. Translate the given text to English accurately."},
                                {"role": "user", "content": prompt}
                            ],
                            max_tokens=150,
                            temperature=0.1
                        )
                    ),
                    timeout=25  # 25 second timeout for OpenAI API call
                )
                
                translated_text = response.choices[0].message.content.strip()
                
                # Track translation token usage
                if response.usage:
                    self.token_usage["translation_tokens"] += response.usage.total_tokens
                    self.token_usage["total_tokens"] += response.usage.total_tokens
                    
                    logger.log_it({
                        "logType": "info",
                        "prefix": "transcription_with_speakers",
                        "logData": {
                            "message": "Translation token usage tracked",
                            "input_tokens": response.usage.prompt_tokens,
                            "output_tokens": response.usage.completion_tokens,
                            "total_tokens": response.usage.total_tokens
                        }
                    })
                
                return {
                    "original_text": text,
                    "translated_text": translated_text,
                    "source_language": source_lang or "unknown",
                    "status": "success"
                }
                
            except asyncio.TimeoutError:
                logger.log_it({
                    "logType": "warning",
                    "prefix": "transcription_with_speakers",
                    "logData": {
                        "message": "OpenAI API timeout during translation",
                        "text": text[:50] + "..." if len(text) > 50 else text
                    }
                })
                return {
                    "original_text": text,
                    "translated_text": text,
                    "source_language": source_lang or "unknown",
                    "status": "timeout"
                }
            
        except Exception as e:
            logger.log_it({
                "logType": "error",
                "prefix": "transcription_with_speakers",
                "logData": {
                    "message": "OpenAI translation failed",
                    "error": str(e),
                    "text": text[:50] + "..." if len(text) > 50 else text
                }
            })
            return {
                "original_text": text,
                "translated_text": text,
                "source_language": source_lang or "unknown",
                "status": "error",
                "error": str(e)
            }
    
    async def process_audio_file_with_speakers(self, audio_path: str, language_override: str = None) -> Dict:
        """
        Complete processing: transcribe with speakers + translate (async version)
        
        Args:
            audio_path: Path to the audio file
            
        Returns:
            dict: Complete processing results with speaker identification
        """
        logger.log_it({
            "logType": "info",
            "prefix": "transcription_with_speakers",
            "logData": f"Processing audio file with speakers: {audio_path}"
        })
        
        try:
            # Step 1: Transcribe with speaker identification (with timeout)
            transcription_result = await asyncio.wait_for(
                asyncio.get_event_loop().run_in_executor(
                    None, self.transcribe_with_speakers, audio_path, language_override
                ),
                timeout=600  # 10 minute timeout for transcription (increased for larger files)
            )
            
            if transcription_result["status"] == "error":
                return transcription_result
            
            # Step 2: Translate segments to English (async with timeouts)
            translated_segments = []
            combined_text = ""
            combined_start = 0
            combined_end = 0
            combined_speaker = "Speaker 1"  # Track speaker for combined segments
            max_segment_length = 15.0
            
            for i, segment in enumerate(transcription_result.get("segments", [])):
                original_text = segment.get("text", "").strip()
                segment_duration = segment.get("end", 0) - segment.get("start", 0)
                current_speaker = segment.get("speaker", "Speaker 1")
                
                # More lenient filtering - only skip completely empty segments
                if not original_text or original_text.strip() == "":
                    continue
                
                # If include_all_segments is True, only combine very short segments
                if self.include_all_segments:
                    # Only combine segments that are extremely short (less than 0.3 seconds) or have very short text (1 character)
                    if len(original_text) < 2 and segment_duration < 0.3:
                        # Add to combined text for better translation
                        if combined_text:
                            combined_text += " " + original_text
                        else:
                            combined_text = original_text
                            combined_start = segment.get("start", 0)
                            combined_speaker = current_speaker
                        combined_end = segment.get("end", 0)
                        continue
                else:
                    # Original logic for backward compatibility
                    if len(original_text) < 2 or segment_duration < 0.5:
                        # Add to combined text for better translation
                        if combined_text:
                            combined_text += " " + original_text
                        else:
                            combined_text = original_text
                            combined_start = segment.get("start", 0)
                            combined_speaker = current_speaker
                        combined_end = segment.get("end", 0)
                        continue
                
                # If we have combined text, translate it first
                if combined_text:
                    try:
                        translation_result = await asyncio.wait_for(
                            self.translate_with_openai(
                                combined_text,
                                transcription_result["detected_language"]
                            ),
                            timeout=30  # 30 second timeout per translation
                        )
                    except asyncio.TimeoutError:
                        logger.log_it({
                            "logType": "warning",
                            "prefix": "transcription_with_speakers",
                            "logData": {
                                "message": "Translation timeout for combined text",
                                "text": combined_text[:100] + "..." if len(combined_text) > 100 else combined_text
                            }
                        })
                        translation_result = {"translated_text": combined_text, "status": "timeout"}
                    except Exception as e:
                        logger.log_it({
                            "logType": "error",
                            "prefix": "transcription_with_speakers",
                            "logData": {
                                "message": "Translation error for combined text",
                                "error": str(e),
                                "text": combined_text[:100] + "..." if len(combined_text) > 100 else combined_text
                            }
                        })
                        translation_result = {"translated_text": combined_text, "status": "error"}
                    
                    # Add combined segment
                    combined_segment = {
                        "start": combined_start,
                        "end": combined_end,
                        "text": combined_text,
                        "confidence": 1.0,
                        "speaker": combined_speaker,
                        "speaker_type": "Agent" if combined_speaker == "Speaker 1" else "Caller",
                        "segment_id": len(translated_segments) + 1,
                        "translated_text": translation_result.get("translated_text", combined_text),
                        "translation_status": translation_result.get("status", "error")
                    }
                    translated_segments.append(combined_segment)
                    combined_text = ""
                
                # Check if current segment is too long and split it
                if segment_duration > max_segment_length:
                    # Split long segment into smaller chunks
                    num_splits = int(segment_duration / max_segment_length) + 1
                    split_duration = segment_duration / num_splits
                    
                    for j in range(num_splits):
                        split_start = segment.get("start", 0) + (j * split_duration)
                        split_end = segment.get("start", 0) + ((j + 1) * split_duration)
                        
                        # For split segments, use original text but adjust timing
                        split_segment = {
                            "start": split_start,
                            "end": split_end,
                            "text": original_text,  # Use full text for each split
                            "confidence": segment.get("confidence", 1.0),
                            "speaker": current_speaker,
                            "speaker_type": "Agent" if current_speaker == "Speaker 1" else "Caller",
                            "segment_id": len(translated_segments) + 1
                        }
                        
                        # Translate split segment with timeout
                        try:
                            translation_result = await asyncio.wait_for(
                                self.translate_with_openai(
                                    original_text,
                                    transcription_result["detected_language"]
                                ),
                                timeout=30  # 30 second timeout per translation
                            )
                        except asyncio.TimeoutError:
                            logger.log_it({
                                "logType": "warning",
                                "prefix": "transcription_with_speakers",
                                "logData": {
                                    "message": "Translation timeout for split segment",
                                    "text": original_text[:100] + "..." if len(original_text) > 100 else original_text
                                }
                            })
                            translation_result = {"translated_text": original_text, "status": "timeout"}
                        except Exception as e:
                            logger.log_it({
                                "logType": "error",
                                "prefix": "transcription_with_speakers",
                                "logData": {
                                    "message": "Translation error for split segment",
                                    "error": str(e),
                                    "text": original_text[:100] + "..." if len(original_text) > 100 else original_text
                                }
                            })
                            translation_result = {"translated_text": original_text, "status": "error"}
                        
                        split_segment.update({
                            "translated_text": translation_result.get("translated_text", original_text),
                            "translation_status": translation_result.get("status", "error")
                        })
                        
                        translated_segments.append(split_segment)
                else:
                    # Translate current segment normally with timeout
                    try:
                        translation_result = await asyncio.wait_for(
                            self.translate_with_openai(
                                original_text,
                                transcription_result["detected_language"]
                            ),
                            timeout=30  # 30 second timeout per translation
                        )
                    except asyncio.TimeoutError:
                        logger.log_it({
                            "logType": "warning",
                            "prefix": "transcription_with_speakers",
                            "logData": {
                                "message": "Translation timeout for segment",
                                "text": original_text[:100] + "..." if len(original_text) > 100 else original_text
                            }
                        })
                        translation_result = {"translated_text": original_text, "status": "timeout"}
                    except Exception as e:
                        logger.log_it({
                            "logType": "error",
                            "prefix": "transcription_with_speakers",
                            "logData": {
                                "message": "Translation error for segment",
                                "error": str(e),
                                "text": original_text[:100] + "..." if len(original_text) > 100 else original_text
                            }
                        })
                        translation_result = {"translated_text": original_text, "status": "error"}
                    
                    translated_segment = {
                        **segment,
                        "translated_text": translation_result.get("translated_text", original_text),
                        "translation_status": translation_result.get("status", "error")
                    }
                    translated_segments.append(translated_segment)
            
            # Handle any remaining combined text
            if combined_text:
                try:
                    translation_result = await asyncio.wait_for(
                        self.translate_with_openai(
                            combined_text,
                            transcription_result["detected_language"]
                        ),
                        timeout=30  # 30 second timeout per translation
                    )
                except asyncio.TimeoutError:
                    logger.log_it({
                        "logType": "warning",
                        "prefix": "transcription_with_speakers",
                        "logData": {
                            "message": "Translation timeout for remaining combined text",
                            "text": combined_text[:100] + "..." if len(combined_text) > 100 else combined_text
                        }
                    })
                    translation_result = {"translated_text": combined_text, "status": "timeout"}
                except Exception as e:
                    logger.log_it({
                        "logType": "error",
                        "prefix": "transcription_with_speakers",
                        "logData": {
                            "message": "Translation error for remaining combined text",
                            "error": str(e),
                            "text": combined_text[:100] + "..." if len(combined_text) > 100 else combined_text
                        }
                    })
                    translation_result = {"translated_text": combined_text, "status": "error"}
                
                combined_segment = {
                    "start": combined_start,
                    "end": combined_end,
                    "text": combined_text,
                    "confidence": 1.0,
                    "speaker": combined_speaker,
                    "speaker_type": "Agent" if combined_speaker == "Speaker 1" else "Caller",
                    "segment_id": len(translated_segments) + 1,
                    "translated_text": translation_result.get("translated_text", combined_text),
                    "translation_status": translation_result.get("status", "error")
                }
                translated_segments.append(combined_segment)
            
            # Combine results
            complete_result = {
                **transcription_result,
                "segments": translated_segments,
                "processing_status": "complete"
            }
            
            return complete_result
            
        except asyncio.TimeoutError:
            logger.log_it({
                "logType": "error",
                "prefix": "transcription_with_speakers",
                "logData": {
                    "message": "Overall processing timeout",
                    "audio_path": audio_path
                }
            })
            return {
                "status": "error",
                "error": "Processing timeout - transcription or translation took too long",
                "audio_path": audio_path
            }
        except Exception as e:
            logger.log_it({
                "logType": "error",
                "prefix": "transcription_with_speakers",
                "logData": {
                    "message": "Processing error",
                    "error": str(e),
                    "audio_path": audio_path
                }
            })
            return {
                "status": "error",
                "error": f"Processing failed: {str(e)}",
                "audio_path": audio_path
            }
    
    def create_excel_with_speakers(self, results_df: pd.DataFrame, output_file: str):
        """
        Create Excel file with transcriptions, speakers, and timestamps
        
        Args:
            results_df: Results dataframe
            output_file: Path to save Excel file
        """
        logger.log_it({
            "logType": "info",
            "prefix": "transcription_with_speakers",
            "logData": f"Creating Excel file with speakers: {output_file}"
        })
        
        # Create a new workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Call Transcriptions with Speakers"
        
        # Define headers
        headers = [
            "File Name",
            "Start Time (seconds)",
            "End Time (seconds)",
            "Duration (seconds)",
            "Speaker",
            "Speaker Type",
            "English Text",
            "Language",
            "Confidence Score",
            "Segment ID"
        ]
        
        # Style for headers
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Write headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border
        
        # Track current row
        current_row = 2
        
        # Process each file
        for _, file_result in results_df.iterrows():
            if file_result['status'] != 'success':
                continue
                
            filename = file_result.get('filename', 'Unknown')
            segments = file_result.get('segments', [])
            
            # If no segments, create one entry for the whole file
            if not segments:
                ws.cell(row=current_row, column=1, value=filename)
                ws.cell(row=current_row, column=2, value=0)
                ws.cell(row=current_row, column=3, value=file_result.get('duration', 0))
                ws.cell(row=current_row, column=4, value=file_result.get('duration', 0))
                ws.cell(row=current_row, column=5, value="Speaker 1")
                ws.cell(row=current_row, column=6, value="Agent")
                ws.cell(row=current_row, column=7, value=file_result.get('translated_text', ''))
                ws.cell(row=current_row, column=8, value=file_result.get('language_name', 'Unknown'))
                ws.cell(row=current_row, column=9, value=1.0)
                ws.cell(row=current_row, column=10, value=1)
                
                # Apply borders
                for col in range(1, 11):
                    ws.cell(row=current_row, column=col).border = border
                
                current_row += 1
            else:
                # Process each segment
                for segment in segments:
                    ws.cell(row=current_row, column=1, value=filename)
                    ws.cell(row=current_row, column=2, value=segment.get('start', 0))
                    ws.cell(row=current_row, column=3, value=segment.get('end', 0))
                    ws.cell(row=current_row, column=4, value=segment.get('end', 0) - segment.get('start', 0))
                    ws.cell(row=current_row, column=5, value=segment.get('speaker', 'Speaker 1'))
                    ws.cell(row=current_row, column=6, value=segment.get('speaker_type', 'Agent'))
                    ws.cell(row=current_row, column=7, value=segment.get('translated_text', segment.get('text', '')))
                    ws.cell(row=current_row, column=8, value=file_result.get('language_name', 'Unknown'))
                    ws.cell(row=current_row, column=9, value=segment.get('confidence', 1.0))
                    ws.cell(row=current_row, column=10, value=segment.get('segment_id', 1))
                    
                    # Apply borders
                    for col in range(1, 11):
                        ws.cell(row=current_row, column=col).border = border
                    
                    current_row += 1
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save the workbook
        wb.save(output_file)
        
        logger.log_it({
            "logType": "info",
            "prefix": "transcription_with_speakers",
            "logData": {
                "message": "Excel file with speakers saved successfully",
                "file_path": output_file,
                "rows_added": current_row - 2
            }
        })
        
        return current_row - 2  # Return number of rows added
    
    def batch_process_with_speakers(self, input_dir: str, output_file: str = None) -> pd.DataFrame:
        """
        Process all audio files in a directory with speaker identification
        
        Args:
            input_dir: Directory containing preprocessed audio files
            output_file: Path to save results CSV (optional)
            
        Returns:
            pd.DataFrame: Results dataframe with speaker information
        """
        input_path = Path(input_dir)
        wav_files = list(input_path.glob("*.wav"))
        
        if not wav_files:
            logger.log_it({
                "logType": "warning",
                "prefix": "transcription_with_speakers",
                "logData": f"No .wav files found in {input_dir}"
            })
            return pd.DataFrame()
        
        logger.log_it({
            "logType": "info",
            "prefix": "transcription_with_speakers",
            "logData": {
                "message": "Starting batch processing with speakers",
                "total_files": len(wav_files),
                "input_directory": input_dir,
                "max_workers": min(3, len(wav_files))
            }
        })
        
        results = []
        successful = 0
        failed = 0
        
        # Use parallel processing for faster execution
        with ThreadPoolExecutor(max_workers=min(3, len(wav_files))) as executor:
            # Submit all tasks
            future_to_file = {
                executor.submit(self.process_audio_file_with_speakers, str(wav_file)): wav_file 
                for wav_file in wav_files
            }
            
            # Process completed tasks with progress tracking
            for i, future in enumerate(as_completed(future_to_file), 1):
                wav_file = future_to_file[future]
                try:
                    result = future.result()
                    result["filename"] = wav_file.name
                    results.append(result)
                    
                    if result.get('status') == 'success':
                        successful += 1
                        logger.log_it({
                            "logType": "info",
                            "prefix": "transcription_with_speakers",
                            "logData": {
                                "message": f"Processed with speakers successfully ({i}/{len(wav_files)})",
                                "filename": wav_file.name,
                                "language": result.get('detected_language', 'unknown'),
                                "speakers": list(result.get('speaker_stats', {}).keys()),
                                "progress": f"{i}/{len(wav_files)} ({i/len(wav_files)*100:.1f}%)"
                            }
                        })
                    else:
                        failed += 1
                        logger.log_it({
                            "logType": "error",
                            "prefix": "transcription_with_speakers",
                            "logData": {
                                "message": f"Processing with speakers failed ({i}/{len(wav_files)})",
                                "filename": wav_file.name,
                                "error": result.get('error', 'Unknown error'),
                                "progress": f"{i}/{len(wav_files)} ({i/len(wav_files)*100:.1f}%)"
                            }
                        })
                        
                except Exception as e:
                    failed += 1
                    logger.log_it({
                        "logType": "error",
                        "prefix": "transcription_with_speakers",
                        "logData": {
                            "message": f"Exception processing file ({i}/{len(wav_files)})",
                            "filename": wav_file.name,
                            "error": str(e),
                            "progress": f"{i}/{len(wav_files)} ({i/len(wav_files)*100:.1f}%)"
                        }
                    })
                    # Add failed result
                    results.append({
                        "file_path": str(wav_file),
                        "filename": wav_file.name,
                        "transcribed_text": "",
                        "translated_text": "",
                        "detected_language": "unknown",
                        "language_name": "Unknown",
                        "duration": 0,
                        "segments": [],
                        "speaker_stats": {},
                        "status": "error",
                        "error": str(e)
                    })
        
        # Convert to DataFrame
        df = pd.DataFrame(results)
        
        # Save results if output file specified
        if output_file:
            df.to_csv(output_file, index=False)
            logger.log_it({
                "logType": "info",
                "prefix": "transcription_with_speakers",
                "logData": f"CSV results with speakers saved to: {output_file}"
            })
        
        logger.log_it({
            "logType": "info",
            "prefix": "transcription_with_speakers",
            "logData": {
                "message": "Batch processing with speakers completed",
                "total_processed": len(results),
                "successful": successful,
                "failed": failed,
                "success_rate": (successful / len(results) * 100) if results else 0
            }
        })
        
        # Log token usage statistics
        token_usage = self.get_token_usage()
        logger.log_it({
            "logType": "info",
            "prefix": "transcription_with_speakers",
            "logData": {
                "message": "Batch processing token usage summary",
                "token_usage": token_usage,
                "files_processed": len(results),
                "successful_files": successful,
                "failed_files": failed
            }
        })
        
        return df
    
    def get_token_usage(self) -> Dict:
        """
        Get current token usage statistics
        
        Returns:
            dict: Token usage summary
        """
        return {
            "transcription_tokens": self.token_usage["transcription_tokens"],
            "translation_tokens": self.token_usage["translation_tokens"],
            "total_tokens": self.token_usage["total_tokens"],
            "estimated_cost_usd": self._estimate_cost()
        }
    
    def reset_token_usage(self):
        """
        Reset token usage tracking
        """
        self.token_usage = {
            "transcription_tokens": 0,
            "translation_tokens": 0,
            "total_tokens": 0
        }
        logger.log_it({
            "logType": "info",
            "prefix": "transcription_with_speakers",
            "logData": {
                "message": "Token usage tracking reset"
            }
        })
    
    def _estimate_cost(self) -> float:
        """
        Estimate cost based on token usage
        
        Returns:
            float: Estimated cost in USD
        """
        # Rough cost estimates (as of 2024)
        # Whisper: ~$0.006 per minute
        # GPT-3.5-turbo: ~$0.002 per 1K tokens
        # GPT-4o-mini: ~$0.00015 per 1K tokens
        
        whisper_cost = 0.0  # Whisper cost is per minute, not per token
        translation_cost = (self.token_usage["translation_tokens"] / 1000) * 0.002  # GPT-3.5-turbo rate
        
        return whisper_cost + translation_cost

    def transcribe_hindi_with_enhancements(self, audio_path: str) -> Dict:
        """
        Enhanced transcription specifically optimized for Hindi audio with Indian code-mixing support
        
        Args:
            audio_path: Path to the audio file
            
        Returns:
            dict: Enhanced transcription results for Hindi
        """
        try:
            logger.log_it({
                "logType": "info",
                "prefix": "transcription_with_speakers",
                "logData": {
                    "message": "Starting enhanced Hindi transcription with Indian code-mixing support",
                    "file": audio_path,
                    "optimizations": self.hindi_optimizations,
                    "use_hinglish": self.use_hinglish
                }
            })
            
            # Check if file exists
            if not os.path.exists(audio_path):
                raise FileNotFoundError(f"Audio file not found: {audio_path}")
            
            # Enhanced prompt for Indian code-mixing with company-specific keywords
            indian_code_mixing_prompt = """
This is a customer service call in India with:
- Hindi as primary language with regional accent (Telugu/Bengali/Marathi/Punjabi/Gujarati)
- English words and phrases mixed in naturally
- Code-switching between Hindi, regional language, and English
- Customer service context with technical terms
- Regional pronunciation variations

IMPORTANT: If X-glish mode is enabled, transcribe Indian languages using English script:
- Telugu → Telugish: "namaskaram" instead of "నమస్కారం"
- Tamil → Tamglish: "vanakkam" instead of "வணக்கம்"  
- Kannada → Kannadish: "namaskara" instead of "ನಮಸ್ಕಾರ"
- Malayalam → Malayglish: "namaskaram" instead of "നമസ്കാരം"
- Bengali → Benglish: "namaskar" instead of "নমস্কার"
- Marathi → Maringlish: "namaskar" instead of "नमस्कार"
- Gujarati → Gujlish: "namaste" instead of "નમસ્તે"
- Punjabi → Punglish: "sat sri akal" instead of "ਸਤ ਸ੍ਰੀ ਅਕਾਲ"
- Odia → Odiaglish: "namaskar" instead of "ନମସ୍କାର"
- Assamese → Asglish: "namaskar" instead of "নমস্কাৰ"

IMPORTANT COMPANY-SPECIFIC KEYWORDS TO RECOGNIZE ACCURATELY:
- Company Names: Yanolja, Yanolja Cloud Solution, Ezee, Ultra Viewer, AnyDesk
- Technical Terms: Cloud Solution,Ultra Viewer, Remote Desktop, AnyDesk, TeamViewer
- Email/Contact: Gmail, Outlook, Yahoo, Hotmail, Mail ID, Email ID
- Common Phrases: "Yanolja Cloud Solution", "Ezee software", "Ultra Viewer", "AnyDesk connection"
- Business Terms: Booking, Reservation, Hotel, Restaurant, Room Charge, Billing
- Technical Support: Remote access, Screen sharing, Connection issues, Login problems
- Hotel Business Terms: Revenue, CityLedger, Front Desk, Check-in, Check-out, Room Rate, Occupancy, RevPAR, ADR, Hotel Management, PMS (Property Management System)
- Financial Terms: Invoice, Payment, Credit Card, Debit Card, Cash, Receipt, Tax, GST, Service Charge, Cancellation Fee, No-show Fee
- Hotel Operations: Housekeeping, Maintenance, Concierge, Bell Boy, Room Service, Housekeeping, Laundry, Mini Bar, Room Key, Key Card
- Revenue Management: Rate Management, Dynamic Pricing, Yield Management, Revenue Optimization, Market Segmentation, Demand Forecasting
- CityLedger Terms: City Ledger, Corporate Account, Direct Billing, Credit Limit, Payment Terms, Aging Report, Outstanding Balance, Credit Application
- Common Agent Requests: Hotel Code, Property Code, AnyDesk Number, Ultra Viewer Password, Mail ID Password, Login Credentials, Access Code, System Password
- Number Recognition Patterns: 
  * Direct numbers: "1, 2, 7, 8" or "one two seven eight"
  * Spelled out: "double nine seven four zero" = "99740"
  * Mixed format: "nine double seven four" = "9774"
  * Repeated digits: "double zero" = "00", "triple five" = "555"
  * Common patterns: "zero zero" = "00", "one one" = "11", "double zero" = "00"
  * Date and time patterns: "12th of June" = "12th June", "1st of January" = "1st January 2025", "last week" = "last week", "next month" = "next month"
- Authentication Terms: Username, Password, Login, Credentials, Access Code, PIN, OTP, Verification Code, Security Code

Please transcribe accurately maintaining:
- Mixed language structure in English script
- Regional accent patterns
- English technical terms
- Natural code-switching
- Customer service terminology
- Proper transliteration of Hindi/regional words
- Accurate recognition of company names and technical terms
- Proper spelling of email addresses and technical terms
"""
            
            # Step 1: Do initial language detection
            logger.log_it({
                "logType": "info",
                "prefix": "transcription_with_speakers",
                "logData": "Step 1: Language detection with Indian code-mixing support"
            })
            
            # Use local Whisper for initial language detection
            # Create optimized prompt for local Whisper (shorter but effective)
            local_prompt = """Customer service call in India with Hindi-English code-mixing. 
            Company names: Yanolja, Ezee, Ultra Viewer, AnyDesk. 
            Technical terms: Cloud Solution, Remote Desktop, Login, Password, Email.
            Hotel terms: Booking, Reservation, Check-in, Check-out, Room Rate, Revenue.
            Transcribe accurately with proper English spelling for technical terms."""
            
            result = self.whisper_model.transcribe(
                audio_path,
                language=None,  # Auto-detect language (matches OpenAI API)
                word_timestamps=True,  # Equivalent to timestamp_granularities=["segment"]
                verbose=False,
                initial_prompt=local_prompt,  # Equivalent to OpenAI's prompt parameter
                temperature=0.0,  # Matches OpenAI API exactly
                beam_size=5,  # OpenAI's default beam search
                best_of=5,  # OpenAI's default best_of
                patience=1.0,  # OpenAI's default patience
                length_penalty=1.0,  # OpenAI's default length penalty
                suppress_tokens=[-1],  # Suppress special tokens like OpenAI
                fp16=False  # Use FP32 for better accuracy (OpenAI uses FP32)
            )
            initial_transcript = {
                "text": result["text"],
                "language": result.get("language", "unknown"),
                "segments": result.get("segments", [])
            }
            
            detected_language = initial_transcript["language"]
            logger.log_it({
                "logType": "info",
                "prefix": "transcription_with_speakers",
                "logData": {
                    "message": "Language detected for enhanced processing with Indian code-mixing",
                    "detected_language": detected_language
                }
            })
            
            # Step 2: Determine transcription approach based on detected language
            if detected_language in ["hi", "hin", "hindi"]:
                # Hindi detected - use Hinglish transcription with enhanced prompts
                transcription_language = "en"  # Use English for Hinglish
                transcription_method = "enhanced_auto_hinglish_indian_code_mixing"
                auto_hinglish_enabled = True
                logger.log_it({
                    "logType": "info",
                    "prefix": "transcription_with_speakers",
                    "logData": "Hindi detected - using enhanced Hinglish transcription with Indian code-mixing"
                })
            elif detected_language in ["en", "english"]:
                # English detected - use English transcription
                transcription_language = "en"
                transcription_method = "enhanced_english"
                auto_hinglish_enabled = False
                logger.log_it({
                    "logType": "info",
                    "prefix": "transcription_with_speakers",
                    "logData": "English detected - using enhanced English transcription"
                })
            else:
                # Other language - use user's preference with Indian code-mixing support
                if self.use_hinglish:
                    transcription_language = "en"
                    transcription_method = "enhanced_user_hinglish_indian_code_mixing"
                    auto_hinglish_enabled = True
                else:
                    transcription_language = None  # Auto-detect
                    transcription_method = "enhanced_standard"
                    auto_hinglish_enabled = False
                logger.log_it({
                    "logType": "info",
                    "prefix": "transcription_with_speakers",
                    "logData": f"Other language detected ({detected_language}) - using {transcription_method}"
                })
            
            # Step 3: Perform transcription with determined approach and enhanced prompts
            try:
                # Use local Whisper for transcription
                result = self.whisper_model.transcribe(
                    audio_path,
                    language=transcription_language if transcription_language != "auto" else None,
                    word_timestamps=True,
                    verbose=False
                )
                transcript = {
                    "text": result["text"],
                    "language": result.get("language", transcription_language),
                    "segments": result.get("segments", [])
                }
                
                # Track transcription usage
                estimated_tokens = len(transcript["text"]) // 4 if transcript["text"] else 0
                self.token_usage["transcription_tokens"] += estimated_tokens
                self.token_usage["total_tokens"] += estimated_tokens
                
                # Process segments with Hindi optimizations
                processed_segments = self._process_hindi_segments(transcript["segments"])
                
                result = {
                    "file_path": audio_path,
                    "transcribed_text": transcript["text"],
                    "detected_language": transcript["language"],
                    "language_name": self.language_map.get(transcript["language"], transcript["language"]),
                    "duration": transcript.get("duration", 0),
                    "segments": processed_segments,
                    "speaker_stats": self._calculate_speaker_stats(processed_segments),
                    "status": "success",
                    "transcription_method": transcription_method,
                    "auto_hinglish_enabled": auto_hinglish_enabled
                }
                
                logger.log_it({
                    "logType": "info",
                    "prefix": "transcription_with_speakers",
                    "logData": {
                        "message": "Enhanced transcription successful",
                        "language": result['detected_language'],
                        "text_length": len(result['transcribed_text']),
                        "method": transcription_method,
                        "auto_hinglish_enabled": auto_hinglish_enabled
                    }
                })
                
                return result
                
            except Exception as e:
                logger.log_it({
                    "logType": "error",
                    "prefix": "transcription_with_speakers",
                    "logData": {
                        "message": "Enhanced transcription failed",
                        "error": str(e),
                        "detected_language": detected_language
                    }
                })
                # Return error result
                return {
                    "file_path": audio_path,
                    "transcribed_text": "",
                    "detected_language": detected_language,
                    "language_name": self.language_map.get(detected_language, detected_language),
                    "duration": 0,
                    "segments": [],
                    "speaker_stats": {},
                    "status": "error",
                    "error": str(e),
                    "transcription_method": transcription_method,
                    "auto_hinglish_enabled": auto_hinglish_enabled
                }
            
        except Exception as e:
            logger.log_it({
                "logType": "error",
                "prefix": "transcription_with_speakers",
                "logData": {
                    "message": "Enhanced Hindi transcription failed",
                    "file": audio_path,
                    "error": str(e)
                }
            })
            return {
                "file_path": audio_path,
                "transcribed_text": "",
                "detected_language": "unknown",
                "language_name": "Unknown",
                "duration": 0,
                "segments": [],
                "speaker_stats": {},
                "status": "error",
                "error": str(e),
                "transcription_method": "failed"
            }
    
    def _process_hindi_segments(self, segments: List) -> List[Dict]:
        """
        Process segments with Hindi-specific optimizations
        
        Args:
            segments: Raw segments from Whisper
            
        Returns:
            List of segments with Hindi-specific processing
        """
        processed_segments = []
        
        # Combine short segments for better Hindi context
        combined_text = ""
        combined_start = 0
        combined_end = 0
        min_length = self.hindi_optimizations.get("min_segment_length", 2.0)
        
        for i, segment in enumerate(segments):
            # Handle both TranscriptionSegment objects and dictionaries
            if hasattr(segment, 'start'):
                start_time = segment.start
                end_time = segment.end
                text = segment.text
                confidence = getattr(segment, 'confidence', 1.0)
            else:
                start_time = segment.get('start', 0)
                end_time = segment.get('end', 0)
                text = segment.get('text', '')
                confidence = segment.get('confidence', 1.0)
            
            segment_duration = end_time - start_time
            
            # Combine short segments for better Hindi context
            if self.hindi_optimizations.get("segment_combining", True) and segment_duration < min_length:
                if combined_text:
                    combined_text += " " + text
                else:
                    combined_text = text
                    combined_start = start_time
                combined_end = end_time
                continue
            
            # If we have combined text, add it as a segment
            if combined_text:
                speaker = self._determine_hindi_speaker(i, combined_text)
                combined_segment = {
                    "start": combined_start,
                    "end": combined_end,
                    "text": combined_text,
                    "confidence": confidence,
                    "speaker": speaker,
                    "speaker_type": "Agent" if speaker == "Speaker 1" else "Caller",
                    "segment_id": len(processed_segments) + 1,
                    "processing": "combined"
                }
                processed_segments.append(combined_segment)
                combined_text = ""
            
            # Add current segment
            speaker = self._determine_hindi_speaker(i, text)
            processed_segment = {
                "start": start_time,
                "end": end_time,
                "text": text,
                "confidence": confidence,
                "speaker": speaker,
                "speaker_type": "Agent" if speaker == "Speaker 1" else "Caller",
                "segment_id": len(processed_segments) + 1,
                "processing": "standard"
            }
            processed_segments.append(processed_segment)
        
        # Handle any remaining combined text
        if combined_text:
            speaker = self._determine_hindi_speaker(len(processed_segments), combined_text)
            combined_segment = {
                "start": combined_start,
                "end": combined_end,
                "text": combined_text,
                "confidence": 1.0,
                "speaker": speaker,
                "speaker_type": "Agent" if speaker == "Speaker 1" else "Caller",
                "segment_id": len(processed_segments) + 1,
                "processing": "combined"
            }
            processed_segments.append(combined_segment)
        
        logger.log_it({
            "logType": "info",
            "prefix": "transcription_with_speakers",
            "logData": {
                "message": "Hindi segments processed",
                "total_segments": len(processed_segments),
                "combined_segments": len([s for s in processed_segments if s.get("processing") == "combined"])
            }
        })
        
        return processed_segments
    
    def _determine_hindi_speaker(self, segment_index: int, text: str) -> str:
        """
        Determine speaker for Hindi segments with enhanced logic
        
        Args:
            segment_index: Index of the segment
            text: Text content of the segment
            
        Returns:
            str: Speaker identifier
        """
        # Enhanced Hindi speaker identification
        if segment_index == 0:
            return "Speaker 1"  # First speaker is usually the agent
        
        # Hindi-specific language patterns
        hindi_agent_words = [
            'नमस्ते', 'धन्यवाद', 'सहायता', 'सेवा', 'मदद', 'कृपया', 'आपका', 'स्वागत',
            'hello', 'thank', 'help', 'service', 'assist', 'please', 'welcome'
        ]
        
        hindi_caller_words = [
            'कैसे', 'क्या', 'कब', 'कहाँ', 'क्यों', 'मुझे', 'मेरा', 'मेरी',
            'how', 'what', 'when', 'where', 'why', 'my', 'me', 'can', 'could'
        ]
        
        # Score-based determination
        agent_score = 0
        caller_score = 0
        
        # Check for Hindi agent words
        for word in hindi_agent_words:
            if word.lower() in text.lower():
                agent_score += 1
        
        # Check for Hindi caller words
        for word in hindi_caller_words:
            if word.lower() in text.lower():
                caller_score += 1
        
        # Professional tone indicators (agent)
        if any(word in text.lower() for word in ['sir', 'madam', 'thank', 'please']):
            agent_score += 1
        
        # Question indicators (caller)
        if any(word in text.lower() for word in ['what', 'how', 'why', 'when', 'where']):
            caller_score += 1
        
        # Length-based analysis (agents typically shorter responses)
        if len(text.split()) < 5:
            agent_score += 1
        elif len(text.split()) > 10:
            caller_score += 1
        
        # Determine speaker based on scores
        if agent_score > caller_score:
            return "Speaker 1"  # Agent
        elif caller_score > agent_score:
            return "Speaker 2"  # Caller
        else:
            # Fallback to alternating pattern
            return "Speaker 1" if segment_index % 2 == 0 else "Speaker 2"

    async def translate_hindi_with_enhancements(self, text: str, source_lang: str = "hi") -> Dict:
        """
        Enhanced translation specifically optimized for Hindi text
        
        Args:
            text: Hindi text to translate
            source_lang: Source language code (default: hi for Hindi)
            
        Returns:
            dict: Enhanced translation results for Hindi
        """
        try:
            logger.log_it({
                "logType": "info",
                "prefix": "transcription_with_speakers",
                "logData": {
                    "message": "Starting enhanced Hindi translation",
                    "text_length": len(text),
                    "source_language": source_lang
                }
            })
            
            if not text.strip() or len(text.strip()) < 3:
                return {
                    "original_text": text,
                    "translated_text": text,
                    "source_language": source_lang,
                    "status": "too_short",
                    "translation_method": "enhanced_hindi"
                }
            
            # Enhanced prompt for Hindi translation
            enhanced_prompt = f"""Translate the following Hindi text to English. 
            Pay special attention to:
            1. Preserve the meaning and context
            2. Handle Hindi-specific expressions and idioms
            3. Maintain the tone and formality level
            4. Ensure proper English grammar and flow
            
            Hindi text: {text}
            
            Provide only the English translation:"""
            
            # Use OpenAI for enhanced Hindi translation
            response = self.client.chat.completions.create(
                model="gpt-4o-mini",  # Use more capable model for Hindi
                messages=[
                    {
                        "role": "system", 
                        "content": "You are an expert Hindi to English translator with deep understanding of both languages, cultural context, and business terminology."
                    },
                    {
                        "role": "user", 
                        "content": enhanced_prompt
                    }
                ],
                max_tokens=300,  # Allow more tokens for Hindi
                temperature=0.1,  # Low temperature for consistent translation
                presence_penalty=0.1,  # Encourage diverse vocabulary
                frequency_penalty=0.1   # Reduce repetition
            )
            
            translated_text = response.choices[0].message.content.strip()
            
            # Track translation token usage
            if response.usage:
                self.token_usage["translation_tokens"] += response.usage.total_tokens
                self.token_usage["total_tokens"] += response.usage.total_tokens
                
                logger.log_it({
                    "logType": "info",
                    "prefix": "transcription_with_speakers",
                    "logData": {
                        "message": "Enhanced Hindi translation token usage tracked",
                        "input_tokens": response.usage.prompt_tokens,
                        "output_tokens": response.usage.completion_tokens,
                        "total_tokens": response.usage.total_tokens
                    }
                })
            
            # Validate Hindi translation quality
            quality_check = self._validate_hindi_translation(text, translated_text)
            
            result = {
                "original_text": text,
                "translated_text": translated_text,
                "source_language": source_lang,
                "status": "success",
                "translation_method": "enhanced_hindi",
                "quality_check": quality_check
            }
            
            logger.log_it({
                "logType": "info",
                "prefix": "transcription_with_speakers",
                "logData": {
                    "message": "Enhanced Hindi translation completed",
                    "original_length": len(text),
                    "translated_length": len(translated_text),
                    "quality_score": quality_check.get("score", 0)
                }
            })
            
            return result
            
        except Exception as e:
            logger.log_it({
                "logType": "error",
                "prefix": "transcription_with_speakers",
                "logData": {
                    "message": "Enhanced Hindi translation failed",
                    "error": str(e),
                    "text": text[:50] + "..." if len(text) > 50 else text
                }
            })
            return {
                "original_text": text,
                "translated_text": text,
                "source_language": source_lang,
                "status": "error",
                "error": str(e),
                "translation_method": "enhanced_hindi_failed"
            }
    
    def _validate_hindi_translation(self, original_text: str, translated_text: str) -> Dict:
        """
        Validate Hindi translation quality
        
        Args:
            original_text: Original Hindi text
            translated_text: Translated English text
            
        Returns:
            dict: Quality validation results
        """
        quality_check = {
            "score": 0,
            "issues": [],
            "strengths": []
        }
        
        # Basic checks
        if not translated_text or translated_text.strip() == "":
            quality_check["issues"].append("Empty translation")
            return quality_check
        
        if translated_text == original_text:
            quality_check["issues"].append("Translation identical to original")
            return quality_check
        
        # Length check (Hindi to English typically results in shorter text)
        length_ratio = len(translated_text) / len(original_text) if len(original_text) > 0 else 0
        if length_ratio < 0.3:
            quality_check["issues"].append("Translation too short")
        elif length_ratio > 2.0:
            quality_check["issues"].append("Translation too long")
        else:
            quality_check["strengths"].append("Appropriate length ratio")
            quality_check["score"] += 20
        
        # Check for common Hindi words that should be translated
        hindi_words = ['नमस्ते', 'धन्यवाद', 'कैसे', 'हैं', 'में', 'का', 'की', 'के', 'है']
        untranslated_hindi = [word for word in hindi_words if word in translated_text]
        if untranslated_hindi:
            quality_check["issues"].append(f"Untranslated Hindi words: {untranslated_hindi}")
        else:
            quality_check["strengths"].append("No Hindi words in translation")
            quality_check["score"] += 30
        
        # Check for English words (good sign)
        english_words = ['the', 'and', 'is', 'are', 'in', 'of', 'to', 'for', 'with']
        english_word_count = sum(1 for word in english_words if word.lower() in translated_text.lower())
        if english_word_count > 2:
            quality_check["strengths"].append("Contains English words")
            quality_check["score"] += 25
        
        # Check for coroutine objects (bad sign)
        if "coroutine" in translated_text.lower():
            quality_check["issues"].append("Translation contains coroutine object")
            quality_check["score"] = 0
        else:
            quality_check["score"] += 25
        
        return quality_check

    def _recognize_number_patterns(self, text: str) -> str:
        """
        Recognize and convert various number patterns to standard format
        """
        import re
        
        # Common number patterns in Indian English/Hinglish
        patterns = {
            r'\b(double)\s+(\d)\b': r'\2\2',  # "double nine" -> "99"
            r'\b(triple)\s+(\d)\b': r'\2\2\2',  # "triple five" -> "555"
            r'\b(zero\s+zero)\b': '00',
            r'\b(one\s+one)\b': '11',
            r'\b(two\s+two)\b': '22',
            r'\b(three\s+three)\b': '33',
            r'\b(four\s+four)\b': '44',
            r'\b(five\s+five)\b': '55',
            r'\b(six\s+six)\b': '66',
            r'\b(seven\s+seven)\b': '77',
            r'\b(eight\s+eight)\b': '88',
            r'\b(nine\s+nine)\b': '99',
        }
        
        processed_text = text
        for pattern, replacement in patterns.items():
            processed_text = re.sub(pattern, replacement, processed_text, flags=re.IGNORECASE)
        
        return processed_text

    async def transcribe_hindi_with_translation(self, audio_path: str) -> Dict:
        """
        Enhanced Hindi transcription with translation to English
        
        Args:
            audio_path: Path to Hindi audio file
            
        Returns:
            dict: Complete Hindi transcription and translation results
        """
        try:
            logger.log_it({
                "logType": "info",
                "prefix": "transcription_with_speakers",
                "logData": f"Starting enhanced Hindi transcription: {audio_path}"
            })
            
            # Step 1: Transcribe with Hindi enhancements
            transcription_result = self.transcribe_hindi_with_enhancements(audio_path)
            
            if transcription_result["status"] == "error":
                return transcription_result
            
            # Check if detected language is English - skip translation
            detected_language = transcription_result.get("detected_language", "").lower()
            if detected_language in ["en", "english"]:
                logger.log_it({
                    "logType": "info",
                    "prefix": "transcription_with_speakers",
                    "logData": {
                        "message": "English audio detected - skipping translation",
                        "detected_language": detected_language
                    }
                })
                
                # Process segments without translation for English audio
                processed_segments = []
                for i, segment in enumerate(transcription_result.get("segments", [])):
                    processed_segment = {
                        **segment,
                        "translated_text": segment.get("text", ""),  # Use original text
                        "translation_status": "no_translation_needed",
                        "translation_method": "none"
                    }
                    processed_segments.append(processed_segment)
                
                final_result = {
                    **transcription_result,
                    "segments": processed_segments,
                    "translated_text": transcription_result.get("transcribed_text", ""),
                    "translation_method": "none"
                }
                
                return final_result
            
            # Step 2: Translate segments to English (only for non-English content)
            translated_segments = []
            combined_text = ""
            combined_start = 0
            combined_end = 0
            combined_speaker = "Speaker 1"  # Track speaker for combined segments
            
            for i, segment in enumerate(transcription_result.get("segments", [])):
                original_text = segment.get("text", "").strip()
                segment_duration = segment.get("end", 0) - segment.get("start", 0)
                current_speaker = segment.get("speaker", "Speaker 1")
                
                # More lenient filtering - only skip completely empty segments
                if not original_text or original_text.strip() == "":
                    continue
                
                # If include_all_segments is True, only combine very short segments
                if self.include_all_segments:
                    # Only combine segments that are extremely short (less than 0.3 seconds) or have very short text (1 character)
                    if len(original_text) < 2 and segment_duration < 0.3:
                        # Add to combined text for better translation
                        if combined_text:
                            combined_text += " " + original_text
                        else:
                            combined_text = original_text
                            combined_start = segment.get("start", 0)
                            combined_speaker = current_speaker
                        combined_end = segment.get("end", 0)
                        continue
                else:
                    # Original logic for backward compatibility
                    if len(original_text) < 2 or segment_duration < 0.5:
                        # Add to combined text for better translation
                        if combined_text:
                            combined_text += " " + original_text
                        else:
                            combined_text = original_text
                            combined_start = segment.get("start", 0)
                            combined_speaker = current_speaker
                        combined_end = segment.get("end", 0)
                        continue
                
                # If we have combined text, translate it first
                if combined_text:
                    translation_result = await self.translate_hindi_with_enhancements(
                        combined_text,
                        transcription_result["detected_language"]
                    )
                    
                    # Add combined segment
                    combined_segment = {
                        "start": combined_start,
                        "end": combined_end,
                        "text": combined_text,
                        "confidence": segment.get("confidence", 1.0),
                        "speaker": combined_speaker,
                        "speaker_type": "Agent" if combined_speaker == "Speaker 1" else "Caller",
                        "segment_id": len(translated_segments) + 1,
                        "translated_text": translation_result.get("translated_text", combined_text),
                        "translation_status": translation_result.get("status", "error"),
                        "translation_method": translation_result.get("translation_method", "enhanced_hindi")
                    }
                    translated_segments.append(combined_segment)
                    combined_text = ""
                
                # Translate current segment
                translation_result = await self.translate_hindi_with_enhancements(
                    original_text,
                    transcription_result["detected_language"]
                )
                
                # Add translated segment
                translated_segment = {
                    "start": segment.get("start", 0),
                    "end": segment.get("end", 0),
                    "text": original_text,
                    "confidence": segment.get("confidence", 1.0),
                    "speaker": current_speaker,
                    "speaker_type": "Agent" if current_speaker == "Speaker 1" else "Caller",
                    "segment_id": len(translated_segments) + 1,
                    "translated_text": translation_result.get("translated_text", original_text),
                    "translation_status": translation_result.get("status", "error"),
                    "translation_method": translation_result.get("translation_method", "enhanced_hindi")
                }
                translated_segments.append(translated_segment)
            
            # Handle any remaining combined text
            if combined_text:
                translation_result = await self.translate_hindi_with_enhancements(
                    combined_text,
                    transcription_result["detected_language"]
                )
                
                combined_segment = {
                    "start": combined_start,
                    "end": combined_end,
                    "text": combined_text,
                    "confidence": 1.0,
                    "speaker": combined_speaker,
                    "speaker_type": "Agent" if combined_speaker == "Speaker 1" else "Caller",
                    "segment_id": len(translated_segments) + 1,
                    "translated_text": translation_result.get("translated_text", combined_text),
                    "translation_status": translation_result.get("status", "error"),
                    "translation_method": translation_result.get("translation_method", "enhanced_hindi")
                }
                translated_segments.append(combined_segment)
            
            # Create final result
            final_result = {
                "file_path": audio_path,
                "transcribed_text": transcription_result.get("transcribed_text", ""),
                "translated_text": " ".join([seg.get("translated_text", "") for seg in translated_segments]),
                "detected_language": transcription_result.get("detected_language", "hi"),
                "language_name": transcription_result.get("language_name", "Hindi"),
                "duration": transcription_result.get("duration", 0),
                "segments": translated_segments,
                "speaker_stats": self._calculate_speaker_stats(translated_segments),
                "status": "success",
                "transcription_method": transcription_result.get("transcription_method", "enhanced_hindi"),
                "translation_method": "enhanced_hindi"
            }
            
            logger.log_it({
                "logType": "info",
                "prefix": "transcription_with_speakers",
                "logData": {
                    "message": "Hindi transcription with translation completed",
                    "file": os.path.basename(audio_path),
                    "segments_count": len(translated_segments),
                    "translation_tokens": self.token_usage.get("translation_tokens", 0)
                }
            })
            
            return final_result
            
        except Exception as e:
            logger.log_it({
                "logType": "error",
                "prefix": "transcription_with_speakers",
                "logData": {
                    "message": "Hindi transcription with translation failed",
                    "file": audio_path,
                    "error": str(e)
                }
            })
            return {
                "file_path": audio_path,
                "transcribed_text": "",
                "translated_text": "",
                "detected_language": "unknown",
                "language_name": "Unknown",
                "duration": 0,
                "segments": [],
                "speaker_stats": {},
                "status": "error",
                "error": str(e),
                "transcription_method": "failed",
                "translation_method": "failed"
            }

    def batch_process_with_advanced_diarization(self, input_dir: str, output_file: str = None) -> pd.DataFrame:
        """
        Process all audio files in a directory with advanced ML-based speaker diarization
        
        Args:
            input_dir: Directory containing preprocessed audio files
            output_file: Path to save results CSV (optional)
            
        Returns:
            pd.DataFrame: Results dataframe with advanced speaker identification
        """
        if not self.advanced_diarization:
            logger.log_it({
                "logType": "warning",
                "prefix": "transcription_with_speakers",
                "logData": "Advanced diarization not initialized, using standard processing"
            })
            return self.batch_process_with_speakers(input_dir, output_file)
        
        input_path = Path(input_dir)
        wav_files = list(input_path.glob("*.wav"))
        
        if not wav_files:
            logger.log_it({
                "logType": "warning",
                "prefix": "transcription_with_speakers",
                "logData": f"No .wav files found in {input_dir}"
            })
            return pd.DataFrame()
        
        logger.log_it({
            "logType": "info",
            "prefix": "transcription_with_speakers",
            "logData": {
                "message": "Starting batch processing with advanced ML diarization",
                "total_files": len(wav_files),
                "input_directory": input_dir,
                "diarization_method": "advanced_ml"
            }
        })
        
        results = []
        successful = 0
        failed = 0
        
        # Process files with advanced diarization
        for i, wav_file in enumerate(wav_files, 1):
            try:
                logger.log_it({
                    "logType": "info",
                    "prefix": "transcription_with_speakers",
                    "logData": {
                        "message": f"Processing with advanced diarization ({i}/{len(wav_files)})",
                        "filename": wav_file.name
                    }
                })
                
                # Load audio for advanced diarization
                audio, sr = librosa.load(str(wav_file), sr=16000)
                
                # Split audio into segments for diarization
                segment_length = int(3 * sr)  # 3-second segments
                audio_segments = []
                
                for j in range(0, len(audio), segment_length):
                    segment = audio[j:j + segment_length]
                    if len(segment) >= segment_length // 2:  # At least 1.5 seconds
                        audio_segments.append(segment)
                
                # Extract features for all segments
                features = []
                for segment in audio_segments:
                    feature_vector = self.advanced_diarization.extract_speaker_features(segment)
                    features.append(feature_vector)
                
                # Cluster speakers using advanced diarization
                speaker_labels = self.advanced_diarization._cluster_speakers(features)
                
                # Now do transcription with speaker labels
                result = self.transcribe_with_speakers(str(wav_file))
                result["filename"] = wav_file.name
                
                # Apply advanced speaker labels to segments
                if result.get("segments"):
                    for k, segment in enumerate(result["segments"]):
                        if k < len(speaker_labels):
                            segment["speaker"] = speaker_labels[k]
                            segment["speaker_type"] = "Agent" if speaker_labels[k] == "Speaker 1" else "Caller"
                            segment["diarization_method"] = "advanced_ml"
                
                results.append(result)
                
                if result.get('status') == 'success':
                    successful += 1
                    logger.log_it({
                        "logType": "info",
                        "prefix": "transcription_with_speakers",
                        "logData": {
                            "message": f"Advanced diarization completed ({i}/{len(wav_files)})",
                            "filename": wav_file.name,
                            "segments": len(result.get("segments", [])),
                            "speaker_distribution": dict(zip(*np.unique(speaker_labels, return_counts=True)))
                        }
                    })
                else:
                    failed += 1
                    
            except Exception as e:
                failed += 1
                logger.log_it({
                    "logType": "error",
                    "prefix": "transcription_with_speakers",
                    "logData": {
                        "message": f"Advanced diarization failed ({i}/{len(wav_files)})",
                        "filename": wav_file.name,
                        "error": str(e)
                    }
                })
                # Add failed result
                results.append({
                    "file_path": str(wav_file),
                    "filename": wav_file.name,
                    "transcribed_text": "",
                    "translated_text": "",
                    "detected_language": "unknown",
                    "language_name": "Unknown",
                    "duration": 0,
                    "segments": [],
                    "speaker_stats": {},
                    "status": "error",
                    "error": str(e),
                    "diarization_method": "advanced_ml_failed"
                })
        
        # Convert to DataFrame
        df = pd.DataFrame(results)
        
        # Save results if output file specified
        if output_file:
            df.to_csv(output_file, index=False)
            logger.log_it({
                "logType": "info",
                "prefix": "transcription_with_speakers",
                "logData": f"Advanced diarization results saved to: {output_file}"
            })
        
        logger.log_it({
            "logType": "info",
            "prefix": "transcription_with_speakers",
            "logData": {
                "message": "Advanced ML diarization batch processing completed",
                "total_processed": len(results),
                "successful": successful,
                "failed": failed,
                "success_rate": (successful / len(results) * 100) if results else 0,
                "diarization_method": "advanced_ml"
            }
        })
        
        return df

def main():
    """Example usage of the TranscriptionWithSpeakersService"""
    # Initialize service
    transcription_service = TranscriptionWithSpeakersService()
    
    # Reset token usage at start
    transcription_service.reset_token_usage()
    
    # Process preprocessed audio files
    input_dir = "Call_recordings_AI/results/preprocessed_audio"
    output_csv = "transcription_results_with_speakers.csv"
    output_excel = "call_transcriptions_with_speakers.xlsx"
    
    # Batch process all files
    results_df = transcription_service.batch_process_with_speakers(input_dir, output_csv)
    
    # Create Excel file with speakers
    if not results_df.empty:
        transcription_service.create_excel_with_speakers(results_df, output_excel)
        
        # Log summary
        successful_count = len(results_df[results_df['status'] == 'success'])
        languages_detected = results_df['detected_language'].value_counts().to_dict()
        
        logger.log_it({
            "logType": "info",
            "prefix": "transcription_with_speakers",
            "logData": {
                "message": "Processing Summary",
                "total_files_processed": len(results_df),
                "successful_transcriptions": successful_count,
                "languages_detected": languages_detected
            }
        })
        
        # Display token usage statistics
        token_usage = transcription_service.get_token_usage()
        
        logger.log_it({
            "logType": "info",
            "prefix": "transcription_with_speakers",
            "logData": {
                "message": "Token Usage Statistics",
                "transcription_tokens": token_usage['transcription_tokens'],
                "translation_tokens": token_usage['translation_tokens'],
                "total_tokens": token_usage['total_tokens'],
                "estimated_cost_usd": token_usage['estimated_cost_usd']
            }
        })
        
        # Log token usage
        logger.log_it({
            "logType": "info",
            "prefix": "transcription_with_speakers",
            "logData": {
                "message": "Processing complete - token usage summary",
                "token_usage": token_usage,
                "files_processed": len(results_df),
                "successful_transcriptions": successful_count
            }
        })

if __name__ == "__main__":
    main() 